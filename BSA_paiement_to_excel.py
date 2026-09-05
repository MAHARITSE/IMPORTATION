# -*- coding: utf-8 -*-
"""
BSA_paiement_to_excel.py — Conversion multi-société
=====================================================
Conversion des relevés de remboursements (PDF) en fichiers Excel
selon le modèle "Modele_Import_Reglements_Decompte_Assurance.xlsx"
(feuille Modele_Reglements, 13 colonnes).

Sociétés supportées : BSA, ARO, SARO, ORANGE, ORANGE MONEY,
AFG ASSURANCES, ALLIANZ, BRED, ACCES, GRT, BMOI, BFV-SG, VIVO, etc.
La société est détectée automatiquement depuis le nom du fichier PDF.

Format traité (propre à BSA) : RELEVE DE REMBOURSEMENTS DES FRAIS DE SANTE
  - ordre de virement en 1re page : N°, Lot, "A , le 17/04/2026",
    "virement de 928 750,00 MGA"
  - puis un bloc par remboursement :
      "1071921-1 ADHESION: 950179 RAKOTOARINAIVO CLOTAIRE CG Client: ..."
      "05/02/2026 RAKOTOARINAIVO ASSOCIATION DISPENSAIRE LOTERANA
       15 000,00 0,00 95,00 14 250,00 750,00 0,00"
      (date | nom patient | executant | FR.REELS | 1ERE MUT | Tx (%) |
       REMB | NON REMB | TPG*)
      Les quatre montants Excel sont déterminés par les règles métier BSA
      selon Tx, FR.REELS, REMB, NON REMB et TPG*.
  - page finale : facture SALFA (n° FA-...) et "Total général".

Utilisation (double-clic sur BSA_paiement.bat, ou invite de commandes) :
    python BSA_paiement_to_excel.py                    # tous les PDF du sous-dossier PDF/
    python BSA_paiement_to_excel.py --force            # régénérer (écrase l'Excel existant)
    python BSA_paiement_to_excel.py "mon_releve.pdf"   # un seul PDF (nom ou chemin)

Sortie : BSA/<ANNEE_REGLEMENT>/<ANNEE_SOINS>/<DATE_PAIEMENT> BSA <ANNEE> <PERIODE> MONTANT <MONTANT>Ar.xlsx
    Deux sous-dossiers, créés automatiquement :
    - <ANNEE_REGLEMENT> = année du virement (le paiement fait CETTE année)
    - <ANNEE_SOINS>     = année de la période de soins payée (1re date de
                          soins), car un paiement de cette année peut régler
                          des soins de l'année dernière
    exemple : BSA/2026/2026/17-04-26 BSA 2026 27-01-26 à 23-02-26 MONTANT 928 750Ar.xlsx
    - DATE_PAIEMENT : date du virement (ligne "A , le 17/04/2026"), au format JJ-MM-AA
    - SOCIETE       : BSA (fixée dans ce script)
    - ANNEE         : année du virement (celle du 1er sous-dossier)
    - PERIODE       : 1re et dernière date de soins du relevé, au format JJ-MM-AA
                      (une seule date si toutes les lignes sont du même jour)
    - MONTANT       : total payé (somme des REMB = montant du virement),
                      précédé du mot "MONTANT" et suivi de "Ar"

Numéro de facture SALFA (colonne Numero_Facture_Prescription) :
  - ancien format : un n° unique "FA-02/BFV/26-022" pour tout le relevé ;
  - format 2025   : un n° par décompte, lu en fin de chaque décompte :
        "Total décompte : 1015497"
        "Date facture: ... FACTURE SALFA TOLIARA N°006-25/BFV/BSA/SA"
    Chaque ligne Excel reçoit le n° de facture de SON décompte.

PDF en erreur : si un PDF est illisible, de format non reconnu, sans ligne
ou sans date, il est DÉPLACÉ dans le sous-dossier BSA/ERREUR/ (créé
automatiquement). Une erreur de configuration ou d'écriture Excel ne déplace
jamais le PDF source. Les PDF du dossier ERREUR ne sont pas retraités :
remettez le PDF dans PDF/ après correction pour réessayer.

Les fichiers Excel déjà existants ne sont PAS écrasés (protection des
modifications manuelles), sauf avec l'option --force.
"""
import re
import sys
import glob
import os
import shutil
import unicodedata
import pdfplumber
import openpyxl.styles
from openpyxl import load_workbook

# Le script se trouve dans le dossier de la société BSA.
# Les PDF à convertir sont déposés dans le sous-dossier "PDF" de ce dossier
# (BSA/PDF/) ; à défaut, ils sont cherchés directement dans le dossier BSA.
# Le modèle Excel est recherché à côté du script, puis dans son dossier parent.
# Les Excel produits sont classés dans un sous-dossier au nom de l'année du
# règlement, puis dans un sous-dossier au nom de l'année des soins payés
# (un paiement de cette année peut régler des soins de l'année dernière) :
# ex : BSA/2026/2026/..., créés automatiquement.
PDF_DIR = os.path.dirname(os.path.abspath(__file__))
PDF_SUBDIR = os.path.join(PDF_DIR, "PDF")

# Le dépôt peut être utilisé directement (script et modèle côte à côte) ou
# avec le script placé dans un sous-dossier de société. On privilégie le
# modèle voisin du script, puis celui du dossier parent pour rester compatible
# avec les deux organisations.
_MODEL_NAME = "Modele_Import_Reglements_Decompte_Assurance.xlsx"
_MODEL_CANDIDATES = [
    os.path.join(PDF_DIR, _MODEL_NAME),
    os.path.join(os.path.dirname(PDF_DIR), _MODEL_NAME),
]
MODEL = next((p for p in _MODEL_CANDIDATES if os.path.isfile(p)),
             _MODEL_CANDIDATES[0])
SHEET = "Modele_Reglements"
SOCIETE = "BSA"  # société par défaut (fallback)

HEADERS = ["Ref_Decompte", "Date_Reglement", "Date_Soins", "Nom_Agent", "Matricule",
           "Numero_Facture_Prescription", "Code_Acte", "Libelle_Acte",
           "Montant_Reclame_Brut", "Ticket_Moderateur", "Montant_Paye_Regle",
           "Montant_Exclu_Rejet", "Motif_Observation"]

# Détection de la société depuis le nom du fichier PDF.
# Formats reconnus :
#   "[2025-07-17] 08-07-2025 ARO 1090270 SALFATU..."
#   "14-08-2026 AFG ASSURANCES 1050265 SALFATU..."
#   "[2025-08-21] 18-08-2025 ORANGE MONEY 1127130 SALFATU..."
FILENAME_SOCIETE_RE = re.compile(
    r"(?:\[\d{4}-\d{2}-\d{2}\]\s+)?"    # [date] optionnel
    r"\d{2}-\d{2}-\d{4}\s+"              # date JJ-MM-AAAA
    r"(.+?)\s+"                           # société (lazy)
    r"\d{6,}\s+"                          # identifiant numérique
    r"SALFATU",                           # marqueur SALFATU
    re.IGNORECASE
)


def detecter_societe(nom_pdf):
    """Détecte la société depuis le nom du fichier PDF.

    Exemples :
        '[2025-07-17] 08-07-2025 ARO 1090270 SALFATU...'  -> 'ARO'
        '14-08-2026 AFG ASSURANCES 1050265 SALFATU...'     -> 'AFG ASSURANCES'
        '[2025-08-21] 18-08-2025 ORANGE MONEY 1127130 ...' -> 'ORANGE MONEY'
    """
    m = FILENAME_SOCIETE_RE.search(nom_pdf)
    if m:
        return m.group(1).strip().upper()
    return SOCIETE  # fallback


def sans_accent(s):
    """'Février' -> 'FEVRIER' (majuscules sans accent, pour les comparaisons)."""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper()


def amount_to_float(s):
    """Montant d'un PDF -> nombre. '' / None / texte non numérique -> 0.0.

    Gère les deux notations rencontrées dans les PDF :
        française : '15 000,00' / '928 750'  -> 15000.0 / 928750.0
        anglaise  : '75,000' / '1,234.56'    -> 75000.0 / 1234.56
    """
    s = re.sub(r"[^\d.,-]", "", str(s or "").replace("\u00a0", " "))
    if not s:
        return 0.0
    virgule, point = s.rfind(","), s.rfind(".")
    dernier = max(virgule, point)
    if dernier >= 0 and s.count(",") + s.count(".") == 1 \
            and len(s) - dernier - 1 == 3:
        # un seul séparateur suivi de 3 chiffres = séparateur de milliers
        s = s.replace(",", "").replace(".", "")
    elif virgule > point:
        s = s.replace(".", "").replace(",", ".")   # virgule = décimales
    else:
        s = s.replace(",", "")                     # point = décimales
    try:
        return float(s)
    except ValueError:
        return 0.0


def fmt_amount(n):
    """928750.0 -> '928 750' ; 1234.5 -> '1 234,5'"""
    if abs(n - round(n)) < 0.005:
        return f"{int(round(n)):,}".replace(",", " ")
    return f"{n:,.1f}".replace(",", " ").replace(".", ",")


def calcul_montants_bsa(tx, fr_reels, remb, non_remb, tpg):
    """Applique les règles métier BSA aux quatre montants exportés.

    Les comparaisons tolèrent un centime d'écart provenant de l'extraction PDF.
    La règle ``REMB = 0 et TPG = 0`` est prioritaire sur les autres règles.
    En dehors des quatre cas définis, les montants BSA sont conservés tels quels
    (TPG comme ticket modérateur et NON REMB comme montant exclu).
    """
    tx_v = amount_to_float(tx)
    fr_v = amount_to_float(fr_reels)
    remb_v = amount_to_float(remb)
    nonremb_v = amount_to_float(non_remb)
    tpg_v = amount_to_float(tpg)
    egal = lambda a, b: abs(a - b) < 0.01

    # Règle 4 (prioritaire, car elle peut aussi satisfaire la règle 3).
    if egal(remb_v, 0) and egal(tpg_v, 0):
        ticket, paye, exclu = 0, 0, fr_v
    # Règle 1 : Tx = 0, FR.REELS = REMB et NON REMB = TPG.
    elif egal(tx_v, 0) and egal(fr_v, remb_v) and egal(nonremb_v, tpg_v):
        ticket, paye, exclu = 0, fr_v, 0
    # Règle 2 : Tx > 0 et FR.REELS = REMB.
    elif tx_v > 0 and egal(fr_v, remb_v):
        ticket, paye, exclu = 0, fr_v, 0
    # Règle 3 : Tx > 0, FR.REELS > REMB et TPG = 0.
    elif tx_v > 0 and fr_v > remb_v and egal(tpg_v, 0):
        ticket, paye, exclu = nonremb_v, remb_v, 0
    else:
        ticket, paye, exclu = tpg_v, remb_v, nonremb_v

    def entier_si_possible(valeur):
        return int(valeur) if valeur == int(valeur) else valeur

    return {
        "Montant_Reclame_Brut": entier_si_possible(fr_v),
        "Ticket_Moderateur": entier_si_possible(ticket),
        "Montant_Paye_Regle": entier_si_possible(paye),
        "Montant_Exclu_Rejet": entier_si_possible(exclu),
    }

def parse_date(d):
    """'02/01/2026' -> '2026-01-02'"""
    dd, mm, yy = d.strip().split("/")
    year = int(yy)
    year += 2000 if year < 100 else 0
    return f"{year:04d}-{int(mm):02d}-{int(dd):02d}"


# --------------------------------------------------------------------------
# Nom du fichier Excel de sortie
#   <DATE_PAIEMENT> <SOCIETE> <ANNEE> <PERIODE> MONTANT <MONTANT>Ar.xlsx
#   exemple : 17-04-26 BSA 2026 27-01-26 à 23-02-26 MONTANT 928 750Ar.xlsx
# Classé dans deux sous-dossiers : année du règlement PUIS année des soins
# (un paiement de cette année peut régler des soins de l'année dernière) :
#   BSA/2026/2026/17-04-26 BSA 2026 27-01-26 à 23-02-26 MONTANT 928 750Ar.xlsx
# --------------------------------------------------------------------------
# Caractères interdits dans un nom de fichier Windows
INVALIDES = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

DATE_ISO = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def date_courte(iso):
    """'2026-01-27' -> '27-01-26' (JJ-MM-AA)."""
    annee, mois, jour = iso.split("-")
    return f"{jour}-{mois}-{annee[2:]}"


def periode_soins(lignes, defaut=None):
    """Période couverte par le paiement : '27-01-26 à 23-02-26'.

    Prend la 1re et la dernière date de soins (colonne Date_Soins) des lignes
    du fichier. Une seule date si toutes les lignes tombent le même jour.
    À défaut (aucune date de soins lisible), la date de règlement.
    """
    dates = sorted(l.get("Date_Soins") or "" for l in lignes)
    dates = [d for d in dates if DATE_ISO.match(d)]
    if not dates:
        if not (defaut and DATE_ISO.match(defaut)):
            return "SANS DATE"
        dates = [defaut]
    debut, fin = dates[0], dates[-1]
    return date_courte(debut) if debut == fin \
        else f"{date_courte(debut)} à {date_courte(fin)}"


def annee_soins(lignes, defaut=None):
    """Année de la période de soins payée par le règlement : '2024', '2025'…

    C'est l'année de la 1re date de soins (colonne Date_Soins) des lignes :
    un paiement fait cette année peut régler des soins de l'année dernière,
    le sous-dossier permet de les distinguer.
    À défaut de date de soins lisible, on reprend l'année du règlement
    (ou 'SANS DATE' s'il n'y a pas non plus de date de règlement).
    """
    dates = sorted(l.get("Date_Soins") or "" for l in lignes)
    dates = [d for d in dates if DATE_ISO.match(d)]
    if not dates:
        if defaut and DATE_ISO.match(defaut):
            dates = [defaut]
        else:
            return "SANS DATE"
    return dates[0].split("-")[0]


def nom_sortie(societe, date_reglement, lignes, montant):
    """Construit le nom du fichier Excel :

        <DATE_PAIEMENT> <SOCIETE> <ANNEE> <PERIODE> MONTANT <MONTANT>Ar.xlsx
        17-04-26 BSA 2026 27-01-26 à 23-02-26 MONTANT 928 750Ar.xlsx

    date_reglement : 'AAAA-MM-JJ' (date du virement / date comptable).
    """
    annee = date_reglement.split("-")[0]
    nom = (f"{date_courte(date_reglement)} {societe} {annee} "
           f"{periode_soins(lignes, date_reglement)} "
           f"MONTANT {fmt_amount(montant)}Ar")
    nom = re.sub(r"\s+", " ", INVALIDES.sub(" ", nom)).strip()
    return nom + ".xlsx"


def dossier_annee(date_reglement, lignes=None):
    """Chemin complet <ANNEE_REGLEMENT>/<ANNEE_SOINS> (créés si absents).

    Les Excel sont classés par année du règlement PUIS par année de la période
    de soins payée (un paiement de cette année peut régler des soins de
    l'année dernière) :
        BSA/2026/2026/17-04-26 BSA 2026 27-01-26 à 23-02-26 MONTANT 928 750Ar.xlsx

    date_reglement : 'AAAA-MM-JJ'.
    lignes : lignes de soins (pour déterminer l'année de soins) ; si absent,
             l'année de soins reprend l'année du règlement.
    """
    annee = date_reglement.split("-")[0]
    soins = annee_soins(lignes, date_reglement) if lignes is not None else annee
    d = os.path.join(PDF_DIR, annee, soins)
    os.makedirs(d, exist_ok=True)
    return d


def group_words(words, tol=2.5):
    """Regroupe les mots d'une page en lignes (par coordonnée verticale).

    Retourne une liste de dictionnaires : {text, x0, words} où words est
    [(x0, x1, texte), ...] trié de gauche à droite.
    """
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    groups = []
    cur, cur_top = [], None
    for w in words:
        if cur and w["top"] - cur_top > tol:
            groups.append(cur)
            cur, cur_top = [], None
        if not cur:
            cur_top = w["top"]
        cur.append(w)
    if cur:
        groups.append(cur)
    out = []
    for g in groups:
        g = sorted(g, key=lambda w: w["x0"])
        out.append({
            "text": " ".join(w["text"] for w in g),
            "x0": g[0]["x0"],
            "top": min(w["top"] for w in g),
            "bottom": max(w["bottom"] for w in g),
            "words": [(w["x0"], w["x1"], w["text"]) for w in g],
        })
    return out



# --------------------------------------------------------------------------
# Format BSA : RELEVE DE REMBOURSEMENTS DES FRAIS DE SANTE
# --------------------------------------------------------------------------
def date_ocr(texte):
    """Normalise une date de soin, même si l'OCR a lu les ``/`` comme 0/1.

    Exemples réellement rencontrés : ``04/04/2025``, ``0410412025`` et
    ``280042025``. La fonction reste volontairement stricte sur la validité du
    jour, du mois et de l'année afin de ne pas prendre un montant pour une date.
    """
    texte = str(texte or "")
    m = re.search(r"(\d{2})\D+(\d{2})\D+(\d{4})", texte)
    candidats = [m.groups()] if m else []
    chiffres = re.sub(r"\D", "", texte)
    if len(chiffres) == 8:
        candidats.append((chiffres[:2], chiffres[2:4], chiffres[4:]))
    elif len(chiffres) == 9:
        # Un séparateur a été reconnu comme un chiffre : 280042025.
        candidats.extend([
            (chiffres[:2], chiffres[3:5], chiffres[5:]),
            (chiffres[:2], chiffres[2:4], chiffres[5:]),
        ])
    elif len(chiffres) == 10:
        # Les deux séparateurs ont été reconnus comme des chiffres :
        # 0410412025 -> 04/04/2025.
        candidats.append((chiffres[:2], chiffres[3:5], chiffres[6:]))

    for jour, mois, annee in candidats:
        try:
            if 1 <= int(jour) <= 31 and 1 <= int(mois) <= 12 \
                    and 2000 <= int(annee) <= 2099:
                return f"{int(jour):02d}/{int(mois):02d}/{int(annee):04d}"
        except (TypeError, ValueError):
            pass
    return None


def montant_ocr(texte):
    """Lit un montant OCR et corrige la virgule décimale souvent supprimée.

    Les relevés sont en ariary entiers. Ainsi ``720000`` dans la colonne REMB
    représente ``7 200,00`` et non 720 000. Les contrôles croisés entre les
    colonnes effectuent ensuite les corrections des rares chiffres manquants.
    """
    brut = str(texte or "").replace("O", "0").replace("o", "0")
    brut = re.sub(r"[^\d.,-]", "", brut)
    if not brut:
        return None
    if not re.search(r"[.,]", brut):
        try:
            valeur = int(re.sub(r"\D", "", brut))
        except ValueError:
            return None
        if len(brut) >= 4 and brut.endswith("00"):
            valeur /= 100
    else:
        valeur = amount_to_float(brut)
    # Tous les montants MGA de ces relevés finissent par ,00. Arrondir élimine
    # aussi les caractères parasites concaténés par certains calques OCR.
    return int(round(valeur))


def taux_ocr(texte):
    """Lit le taux OCR (``8000`` -> 80, ``900``/``9000`` -> 90)."""
    brut = str(texte or "").replace("O", "0").replace("o", "0")
    chiffres = re.sub(r"\D", "", brut)
    if not chiffres:
        return None
    if re.search(r"[.,]", brut):
        valeur = amount_to_float(brut)
    else:
        valeur = int(chiffres)
        if valeur == 900:             # OCR de 90,00 avec un zéro perdu
            valeur = 90
        elif 1000 <= valeur <= 10000 and valeur % 100 == 0:
            valeur /= 100             # 8000 / 9000 / 10000
    return int(round(valeur))


def _centre(mot):
    return (mot["x0"] + mot["x1"]) / 2


def _texte_mots(mots):
    """Assemble et nettoie une liste de mots pdfplumber."""
    texte = " ".join(w["text"] for w in sorted(
        mots, key=lambda w: (w["top"], w["x0"])))
    texte = re.sub(r"\s+", " ", texte).strip(" |_—–-")
    return texte


COLONNES_MONTANTS = (
    (300, 360),  # FR.REELS
    (360, 405),  # 1ERE MUT
    (405, 445),  # Tx (%)
    (445, 495),  # REMB
    (495, 545),  # NON REMB
    (545, 590),  # TPG*
)


def _valeurs_ligne_ocr(mots_page, mot_date):
    """Extrait les six colonnes chiffrées grâce à leurs coordonnées.

    Une expression régulière sur le texte complet n'est pas assez robuste :
    le calque OCR coupe parfois ``25 000,00`` en deux mots, décale les montants
    de quelques pixels ou supprime une valeur. Les colonnes du PDF, elles,
    restent toujours aux mêmes abscisses.
    """
    proches = [w for w in mots_page
               if 300 <= _centre(w) < 590
               and re.search(r"\d", w["text"])
               and abs(w["bottom"] - mot_date["bottom"]) < 9]
    if not proches:
        return [""] * 6
    # La colonne 1ERE MUT contient presque toujours 0,00 et fournit donc le
    # meilleur alignement. À défaut, n'importe quelle autre colonne chiffrée
    # de la même ligne sert de repère.
    candidats_mutuelle = [w for w in proches if 360 <= _centre(w) < 405]
    candidats_base = candidats_mutuelle or proches
    base = min(candidats_base,
               key=lambda w: abs(w["bottom"] - mot_date["bottom"]))["bottom"]
    valeurs = []
    for gauche, droite in COLONNES_MONTANTS:
        mots = [w for w in mots_page
                if gauche <= _centre(w) < droite
                and re.search(r"\d", w["text"])
                and abs(w["bottom"] - base) <= 1.7]
        valeurs.append("".join(w["text"] for w in sorted(mots,
                                                           key=lambda w: w["x0"])))
    return valeurs


def _score_taux(taux, rembourse, brut):
    """Score d'une hypothèse de correction (plus petit = meilleur)."""
    if not brut or taux is None:
        return 0
    ratio = 100 * rembourse / brut
    if taux == 0:
        return 0 if rembourse == 0 else 100 + ratio
    # Un acte plafonné peut avoir un ratio inférieur au taux contractuel
    # (ex. taux 95 %, remboursement plafonné à 15 000 sur 20 000). En revanche
    # un ratio nettement supérieur au taux lu indique presque toujours une
    # erreur OCR.
    return (ratio - taux) * 4 if ratio > taux else (taux - ratio) * .15


def corriger_montants_ocr(fr, mutuelle, taux, rembourse, non_remb, tpg):
    """Réconcilie les montants d'une ligne à partir des règles du relevé.

    Identité utilisée :
      FR.REELS = 1ERE MUT + REMB + (NON REMB - TPG*)

    Si un seul chiffre est absent ou erroné, le taux permet de choisir la
    colonne à corriger sans modifier les cas métier particuliers (plafond,
    exclusion et tiers payant).
    """
    mutuelle = 0 if mutuelle is None else mutuelle
    tpg = 0 if tpg is None else tpg

    if rembourse is None and fr is not None and non_remb is not None:
        rembourse = max(0, fr - mutuelle - non_remb + tpg)
    if non_remb is None and fr is not None and rembourse is not None:
        non_remb = max(0, fr - mutuelle - rembourse + tpg)
    if rembourse is None or non_remb is None:
        return None

    attendu = mutuelle + rembourse + max(0, non_remb - tpg)
    fr_manquant = fr is None
    if fr_manquant:
        fr = attendu
        # Si REMB a perdu un zéro, la simple somme reste cohérente mais son
        # ratio est impossible (ex. 560 au lieu de 5 600 à 80 %).
        if taux and 0 < taux < 100 and tpg == 0 and mutuelle == 0 and fr:
            ratio = 100 * rembourse / fr
            if ratio > taux + 3 or ratio < taux - 25:
                rembourse_calc = round(non_remb * taux / (100 - taux))
                if rembourse_calc >= 0:
                    rembourse = rembourse_calc
                    fr = rembourse + non_remb
    elif abs(fr - attendu) >= 1:
        if fr <= mutuelle + rembourse and attendu > fr:
            # FR.REELS illisible (0, fragments superposés...) : les autres
            # colonnes permettent de le reconstruire directement.
            fr = attendu
        else:
            # Trois hypothèses, chacune ne corrige qu'une colonne. Le taux lu
            # départage les cas où l'OCR a supprimé un zéro.
            hypotheses = [
                (attendu, rembourse, non_remb),
                (fr, max(0, fr - mutuelle - non_remb + tpg), non_remb),
                (fr, rembourse, max(0, fr - mutuelle - rembourse + tpg)),
            ]
            fr, rembourse, non_remb = min(
                hypotheses,
                key=lambda h: _score_taux(taux, h[1], h[0]))

    # Corrige un taux manifestement inférieur au ratio payé. Les taux
    # contractuels supérieurs au ratio sont conservés (actes plafonnés).
    if fr and rembourse and taux is not None:
        ratio = 100 * rembourse / fr
        taux_valides = (20, 50, 60, 70, 75, 80, 85, 90, 95, 100)
        if taux not in (0, *taux_valides) or ratio > taux + 3:
            taux = min(taux_valides, key=lambda v: abs(v - ratio))

    return tuple(int(round(v)) for v in
                 (fr, mutuelle, taux or 0, rembourse, non_remb, tpg))


def _normaliser_code(code):
    code = re.sub(r"[^A-Za-z0-9€&]", "", code or "").upper()
    corrections = {
        "CE": "CG", "€": "EB", "€8": "EB", "&": "EB",
        "DE": "DC", "DSO": "DSO",
    }
    return corrections.get(code, code)


def _numero_facture(texte):
    """Extrait un numéro de facture, ancien (FA-...) ou nouveau (004-25/...)."""
    normalise = sans_accent(texte)
    ancien = re.search(r"\bFA-\d{2}[-/][A-Z0-9/.-]*\d\b", normalise)
    if ancien:
        return ancien.group(0)
    m = re.search(
        r"(?:(N)\s*[°ºO]\s*)?(\d{3,4}[-/]\d{2}(?:/[A-Z0-9.*-]+)+)",
        normalise)
    if not m:
        return ""
    numero = m.group(2).rstrip(".-")
    return ("N°" if m.group(1) else "") + numero


def parse_bsa(pdf, nom_pdf):
    """Extrait les métadonnées et toutes les lignes, y compris d'un PDF OCR.

    Le parseur s'appuie sur les coordonnées fixes des colonnes plutôt que sur
    une ligne de texte parfaite. Il supporte donc les coupures de milliers,
    virgules supprimées, dates dont les barres sont lues comme 0/1, valeurs
    verticalement décalées et références dont le tiret a disparu.
    """
    pages = []
    lignes_visuelles = []
    for numero_page, page in enumerate(pdf.pages):
        mots = page.extract_words()
        pages.append(mots)
        for ligne in group_words(mots):
            ligne["page"] = numero_page
            lignes_visuelles.append(ligne)
    text = "\n".join(l["text"] for l in lignes_visuelles)

    meta = {"ref": None, "lot": None, "date_reglement": None,
            "virement": None, "facture": None,
            "factures_decompte": {}, "nb_declare": None}
    m = re.search(r"N°\s*:\s*(\d+)", text)
    if m:
        meta["ref"] = m.group(1)
    m = re.search(r"Lot\s*:\s*(\d+)", text)
    if m:
        meta["lot"] = m.group(1)
    m = re.search(r"\bA\s*,?\s*le\s+(\d{2}/\d{2}/\d{4})", text,
                  re.IGNORECASE)
    if m:
        meta["date_reglement"] = m.group(1)
    m = re.search(r"virement de ([\d\u00a0 .]+),(\d{2})\s*MGA", text,
                  re.IGNORECASE)
    if m:
        meta["virement"] = amount_to_float(m.group(1) + "," + m.group(2))
    m = re.search(r"Total g[ée]n[ée]ral\s*:\s*(\d+)\s+(\d+)", text,
                  re.IGNORECASE)
    if m:
        meta["nb_declare"] = int(m.group(2))

    lignes_brutes = []
    decompte_courant = None
    for numero_page, mots_page in enumerate(pages):
        # Les références de bloc se situent dans la première colonne et juste
        # au-dessus de chaque date de soin.
        entetes = [w for w in mots_page
                   if w["x0"] < 90
                   and len(re.sub(r"\D", "", w["text"])) >= 7
                   and not date_ocr(w["text"])]
        dates = [w for w in mots_page
                 if w["x0"] < 50 and date_ocr(w["text"])]
        entetes.sort(key=lambda w: w["top"])
        dates.sort(key=lambda w: w["top"])

        for mot_date in dates:
            precedents = [w for w in entetes
                          if 3 < mot_date["top"] - w["top"] < 30]
            if not precedents:
                print(f"!! {nom_pdf} : date {mot_date['text']} sans bloc -> ignorée")
                continue
            entete = max(precedents, key=lambda w: w["top"])
            suivants = [w["top"] for w in entetes if w["top"] > entete["top"]]
            fin = min(suivants) if suivants else mot_date["top"] + 35

            mots_entete = [w for w in mots_page
                           if abs(w["top"] - entete["top"]) <= 2.5]
            brut_ref = entete["text"]
            chiffres_ref = re.sub(r"\D", "", brut_ref)
            m_ref = re.match(r"(\d{6,})[-–—](\d+)", brut_ref)
            if m_ref:
                decompte_courant = m_ref.group(1)
                numero_ligne = f"{m_ref.group(1)}-{m_ref.group(2)}"
            elif decompte_courant and chiffres_ref.startswith(decompte_courant):
                suite = chiffres_ref[len(decompte_courant):]
                numero_ligne = f"{decompte_courant}-{suite}" if suite else brut_ref
            else:
                numero_ligne = brut_ref

            matricule = _texte_mots([
                w for w in mots_entete
                if 95 <= _centre(w) < 140 and re.search(r"\d", w["text"])
            ])
            code = _normaliser_code(_texte_mots([
                w for w in mots_entete if 245 <= _centre(w) < 290
            ]))

            valeurs = _valeurs_ligne_ocr(mots_page, mot_date)
            fr = montant_ocr(valeurs[0])
            mutuelle = montant_ocr(valeurs[1])
            taux = taux_ocr(valeurs[2])
            rembourse = montant_ocr(valeurs[3])
            non_remb = montant_ocr(valeurs[4])
            tpg = montant_ocr(valeurs[5])
            montants = corriger_montants_ocr(
                fr, mutuelle, taux, rembourse, non_remb, tpg)
            if montants is None:
                print(f"!! {nom_pdf} : montants illisibles pour le bloc "
                      f"{numero_ligne} -> ignoré")
                continue
            fr, mutuelle, taux, rembourse, non_remb, tpg = montants

            base_bas = min(
                (w["bottom"] for w in mots_page
                 if 360 <= _centre(w) < 405
                 and re.search(r"\d", w["text"])
                 and abs(w["bottom"] - mot_date["bottom"]) < 9),
                default=mot_date["bottom"])
            mots_nom = [w for w in mots_page
                        if 45 <= _centre(w) < 132
                        and mot_date["top"] - 2 <= w["top"] < fin]
            nom = _texte_mots([
                w for w in mots_nom
                if re.search(r"[A-Za-zÀ-ÿ]", w["text"])
            ])
            mots_detail = [w for w in mots_page
                           if base_bas + 1 < w["top"] < fin]
            mots_libelle = []
            for ligne_detail in group_words(mots_detail):
                mots_de_ligne = [w for w in mots_detail
                                 if abs(w["top"] - ligne_detail["top"]) <= 2.5
                                 and 132 <= _centre(w) < 590]
                if any(re.search(r"[A-Za-zÀ-ÿ]", w["text"])
                       for w in mots_de_ligne):
                    # Une fois la ligne identifiée comme libellé, conserver
                    # aussi ses nombres (dosage, plafond, n° de facture...).
                    mots_libelle.extend(mots_de_ligne)
            libelle = _texte_mots(mots_libelle)

            motif = f"Prise en charge : {taux:g}%"
            if mutuelle:
                motif += f" ; 1ère mutuelle : {fmt_amount(mutuelle)} Ar"

            montants_excel = calcul_montants_bsa(
                taux, fr, rembourse, non_remb, tpg)
            lignes_brutes.append({
                "page": numero_page,
                "top": entete["top"],
                "decompte": decompte_courant,
                "numero_ligne": numero_ligne,
                "Date_Soins": parse_date(date_ocr(mot_date["text"])),
                "Nom_Agent": nom,
                "Matricule": matricule,
                "Code_Acte": code,
                "Libelle_Acte": libelle,
                "Montant_Reclame_Brut": fr,
                "Ticket_Moderateur": montants_excel["Ticket_Moderateur"],
                "Montant_Paye_Regle": montants_excel["Montant_Paye_Regle"],
                "Montant_Exclu_Rejet": montants_excel["Montant_Exclu_Rejet"],
                "Motif_Observation": motif,
            })

    # La ligne « Date facture » suit toujours les soins de son décompte. Cette
    # position permet l'association même lorsque « Total décompte » a perdu son
    # numéro pendant l'OCR.
    for ligne in lignes_visuelles:
        if not ligne["text"].lower().startswith("date facture"):
            continue
        facture = _numero_facture(ligne["text"])
        if not facture:
            continue
        avant = [l for l in lignes_brutes
                 if (l["page"], l["top"]) < (ligne["page"], ligne["top"])]
        if avant:
            decompte = max(avant, key=lambda l: (l["page"], l["top"]))["decompte"]
            if decompte:
                meta["factures_decompte"][decompte] = facture
        if meta["facture"] is None:
            meta["facture"] = facture

    # Ancien format : un unique numéro FA-... peut apparaître ailleurs que sur
    # une ligne « Date facture ». Il s'applique alors à tout le relevé.
    if meta["facture"] is None:
        meta["facture"] = _numero_facture(text) or None

    lignes = []
    date_reglement = (parse_date(meta["date_reglement"])
                      if meta["date_reglement"] else "")
    for l in lignes_brutes:
        decompte = l.pop("decompte")
        l.pop("page")
        l.pop("top")
        l.pop("numero_ligne")
        lignes.append({
            "Ref_Decompte": meta["ref"] or "",
            "Date_Reglement": date_reglement,
            **l,
            "Numero_Facture_Prescription": (
                meta["factures_decompte"].get(decompte)
                or meta["facture"] or ""),
        })
        # Réordonner selon HEADERS est fait au moment de l'écriture.
    return meta, lignes

# --------------------------------------------------------------------------
# Écriture Excel (mise en forme du modèle)
# --------------------------------------------------------------------------
def style_sheet(ws):
    for r in range(1, ws.max_row + 1):
        for c in range(1, 14):
            ws.cell(row=r, column=c).font = openpyxl.styles.Font(
                name="Calibri", size=12, bold=(r == 1))
    for r in range(2, ws.max_row + 1):
        for c in (9, 10, 11, 12):  # montants
            ws.cell(row=r, column=c).number_format = "#,##0"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"


def write_workbook(path, lignes):
    """Remplit une copie du modèle et l'enregistre de façon atomique.

    Les trois exemples de la feuille ``Modele_Reglements`` sont supprimés,
    tandis que la feuille ``Guide_Rattachement`` et les largeurs de colonnes du
    modèle sont conservées.
    """
    if not os.path.isfile(MODEL):
        raise FileNotFoundError(f"Modèle Excel introuvable : {MODEL}")
    wb = load_workbook(MODEL)
    if SHEET not in wb.sheetnames:
        raise KeyError(f"Feuille {SHEET!r} absente du modèle Excel")
    ws = wb[SHEET]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for colonne, entete in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=colonne, value=entete)
    for ligne in lignes:
        ws.append([ligne[h] for h in HEADERS])
    style_sheet(ws)

    # Évite de laisser un fichier .xlsx incomplet si Excel, le disque ou une
    # synchronisation cloud interrompt la sauvegarde.
    temporaire = path + ".tmp.xlsx"
    try:
        wb.save(temporaire)
        os.replace(temporaire, path)
    finally:
        wb.close()
        if os.path.exists(temporaire):
            os.remove(temporaire)


# --------------------------------------------------------------------------
# Dossier des PDF en échec : BSA/ERREUR (créé automatiquement si besoin).
# Tout PDF impossible à convertir y est DÉPLACÉ, pour ne pas le confondre
# avec les PDF en attente de conversion. Les PDF du dossier ERREUR ne sont
# jamais retraités par les conversions suivantes : après correction, on
# remet le PDF dans PDF/ pour réessayer.
# --------------------------------------------------------------------------
ERREUR_DIRNAME = "ERREUR"
ERREURS = []        # [(nom du PDF, raison)] — récapitulatif en fin de traitement


def dossier_erreur():
    """Chemin du sous-dossier BSA/ERREUR (créé si absent)."""
    d = os.path.join(PDF_DIR, ERREUR_DIRNAME)
    os.makedirs(d, exist_ok=True)
    return d


def est_dans_erreur(pdf_path):
    """Vrai si le PDF se trouve déjà dans le sous-dossier ERREUR."""
    parties = os.path.normpath(os.path.abspath(pdf_path)).split(os.sep)
    return ERREUR_DIRNAME in parties


def deplacer_pdf_en_erreur(pdf_path, raison):
    """Déplace un PDF impossible à convertir dans BSA/ERREUR.

    Ne bouge pas le PDF si le déplacement échoue (fichier verrouillé...) :
    le problème est affiché et le PDF sera retraité au prochain lancement.
    """
    dest = os.path.join(dossier_erreur(), os.path.basename(pdf_path))
    base, ext = os.path.splitext(dest)
    n = 1
    while os.path.exists(dest):     # ne pas écraser un PDF déjà signalé
        dest = f"{base} ({n}){ext}"
        n += 1
    try:
        shutil.move(pdf_path, dest)
        print(f"   -> PDF déplacé dans {ERREUR_DIRNAME} : "
              f"{os.path.relpath(dest, PDF_DIR)}")
        ERREURS.append((os.path.basename(pdf_path), raison))
    except Exception as e:
        print(f"   !! déplacement impossible vers {ERREUR_DIRNAME} : {e}")
        ERREURS.append((os.path.basename(pdf_path), raison + f" (non déplacé : {e})"))


def recap_erreurs():
    """Affiche le récapitulatif des erreurs de traitement."""
    if ERREURS:
        print(f"\n== {len(ERREURS)} erreur(s) de traitement ==")
        for nom, raison in ERREURS:
            print(f"   - {nom} : {raison}")


# --------------------------------------------------------------------------
def main():
    if not os.path.isfile(MODEL):
        print("!! Modèle Excel introuvable. Emplacements vérifiés :")
        for candidat in _MODEL_CANDIDATES:
            print(f"   - {candidat}")
        print("   Les PDF sont conservés dans leur dossier.")
        return 1

    args = [a for a in sys.argv[1:] if a != "--force"]
    force = "--force" in sys.argv[1:]

    # PDF à convertir : ceux passés en argument, sinon tous les PDF du dossier.
    pdfs = []
    if args:
        for a in args:
            if os.path.isabs(a):
                p = a
            else:
                # nom relatif : d'abord dans le sous-dossier PDF/, sinon
                # directement dans le dossier de la société
                p = os.path.join(PDF_SUBDIR, a)
                if not os.path.isfile(p):
                    p = os.path.join(PDF_DIR, a)
            if os.path.isfile(p) and p.lower().endswith(".pdf"):
                pdfs.append(os.path.abspath(p))
            else:
                print(f"!! PDF introuvable : {a}")
    else:
        # PDF déposés dans le sous-dossier PDF/ ; à défaut, directement dans
        # le dossier de la société, puis recherche récursive.
        pdfs = sorted(glob.glob(os.path.join(PDF_SUBDIR, "*.pdf")))
        if not pdfs:
            pdfs = sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf")))
        if not pdfs:
            pdfs = sorted(glob.glob(os.path.join(PDF_DIR, "**", "*.pdf"), recursive=True))
    if not pdfs:
        if args:
            print("!! Aucun des PDF demandés n'a été trouvé")
        else:
            print(f"!! Aucun PDF trouvé dans {PDF_SUBDIR}")
        recap_erreurs()
        return
    # Ne pas retraiter les PDF déjà mis de côté dans ERREUR
    pdfs = [p for p in pdfs if not est_dans_erreur(p)]

    for pdf_path in pdfs:
        nom_pdf = os.path.basename(pdf_path)
        try:
            pdf = pdfplumber.open(pdf_path)
        except Exception as e:
            print(f"!! {nom_pdf} : PDF illisible ({e})")
            deplacer_pdf_en_erreur(pdf_path, "PDF illisible ou corrompu")
            continue
        try:
            with pdf:
                # Dossier BSA → parseur BSA. Le contenu du PDF ne change rien.
                meta, lignes = parse_bsa(pdf, nom_pdf)
        except Exception as e:
            print(f"!! {nom_pdf} : erreur pendant la lecture du PDF "
                  f"({type(e).__name__}: {e})")
            deplacer_pdf_en_erreur(pdf_path, "erreur de lecture (format non reconnu ?)")
            continue

        if not lignes:
            print(f"!! {nom_pdf} : aucune ligne de règlement trouvée "
                  f"(format non reconnu ?)")
            deplacer_pdf_en_erreur(pdf_path, "aucune ligne trouvée dans le PDF")
            continue

        # --- Nom du fichier : DATE_PAIEMENT BSA ANNEE PERIODE MONTANT <montant>Ar,
        #     classé dans <ANNEE du règlement>/<ANNEE des soins> ---
        dr = (meta.get("date_reglement") or "").strip()
        if not re.fullmatch(r"\d{2}/\d{2}/\d{4}", dr):
            print(f"!! {nom_pdf} : date de règlement introuvable dans le PDF")
            deplacer_pdf_en_erreur(pdf_path, "date de règlement introuvable")
            continue
        date_reglement = parse_date(dr)          # 'AAAA-MM-JJ'

        total_paye = sum(l["Montant_Paye_Regle"] for l in lignes)
        ref = f"relevé N° {meta['ref']}" if meta["ref"] else "relevé"
        if meta["virement"] is not None:
            ok = abs(total_paye - meta["virement"]) < 1
            if ok:
                print(f"   contrôle : {fmt_amount(total_paye)} Ar payés "
                      f"= montant du virement ({fmt_amount(meta['virement'])} Ar)  OK")
            else:
                print(f"   !! ATTENTION : {fmt_amount(total_paye)} Ar payés "
                      f"≠ montant du virement ({fmt_amount(meta['virement'])} Ar)")
        if meta["nb_declare"] is not None and meta["nb_declare"] != len(lignes):
            print(f"   !! ATTENTION : {len(lignes)} lignes lues, "
                  f"{meta['nb_declare']} déclarées dans le total général du relevé")

        societe = detecter_societe(nom_pdf)
        out = os.path.join(dossier_annee(date_reglement, lignes),
                           nom_sortie(societe, date_reglement, lignes, total_paye))
        relatif = os.path.relpath(out, PDF_DIR)
        if os.path.exists(out) and not force:
            print(f"-- {relatif} : existe déjà, non écrasé "
                  f"(--force pour régénérer)  [{nom_pdf}]")
            continue
        try:
            write_workbook(out, lignes)
        except Exception as e:
            print(f"!! {nom_pdf} : erreur pendant la création de l'Excel "
                  f"({type(e).__name__}: {e})")
            print("   -> PDF conservé : l'erreur ne vient pas nécessairement du PDF")
            ERREURS.append((nom_pdf, "erreur de création de l'Excel (PDF conservé)"))
            continue
        print(f"OK {relatif} : {ref} | {len(lignes)} lignes | "
              f"Payé {fmt_amount(total_paye)} Ar  <- {nom_pdf}")

    recap_erreurs()
    return 1 if ERREURS else 0


if __name__ == "__main__":
    sys.exit(main())
