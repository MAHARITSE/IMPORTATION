# -*- coding: utf-8 -*-
"""
Conversion des factures PDF SALFA (BSA, MCI, ...) en fichiers Excel selon le modèle
"Modele_Import.xlsx" (feuille Modele_Prestations, 11 colonnes).

Usage :
    python3 pdf_to_excel.py                # traite tous les clients trouvés (BSA, MCI, ...)
    python3 pdf_to_excel.py MCI            # traite uniquement MCI
    python3 pdf_to_excel.py MCI --force    # régénère même si le fichier Excel existe déjà

Entrée : FACTURE CLIENT/<CLIENT> <MOIS>.pdf
Sortie : FACTURE CLIENT/EXCEL/<CLIENT> <MOIS>.xlsx  +  fichier consolidé par client

Les fichiers Excel déjà existants ne sont PAS écrasés (protection des modifications
manuelles), sauf avec l'option --force.
"""
import re
import sys
import glob
import os
import unicodedata
import pdfplumber
import openpyxl.styles
from openpyxl import Workbook, load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
PDF_DIR = os.path.join(BASE, "FACTURE CLIENT")
OUT_DIR = os.path.join(PDF_DIR, "EXCEL")
MODEL = os.path.join(PDF_DIR, "Modele_Import.xlsx")

HEADERS = ["Numero_Facture", "Date_Soins", "Nom_Agent", "Matricule", "Societe",
           "Sous_Societe", "Acte_Medicale_Prix", "Montant_Total_Brut",
           "Ticket_Moderateur", "Prise_En_Charge_Net", "Observations"]

MONTH_ORDER = ["JANVIER", "FEVRIER", "MARS", "AVRIL", "MAI", "JUIN",
               "JUILLET", "AOUT", "SEPTEMBRE", "OCTOBRE", "NOVEMBRE", "DECEMBRE"]


def amount_to_float(s):
    """'1 234,56' -> 1234.56"""
    return float(s.replace("\u00a0", " ").replace(" ", "").replace(",", "."))


def fmt_amount(n):
    """1234.0 -> '1 234' ; 1234.5 -> '1 234,5'"""
    if abs(n - round(n)) < 0.005:
        return f"{int(round(n)):,}".replace(",", " ")
    return f"{n:,.1f}".replace(",", " ").replace(".", ",")


def parse_date(d):
    """'02/01/26' -> '2026-01-02'"""
    dd, mm, yy = d.strip().split("/")
    year = int(yy)
    year += 2000 if year < 100 else 0
    return f"{year:04d}-{mm}-{dd}"


def parse_pdf(path):
    """Extrait (facture, mois, lignes) d'un PDF. Chaque ligne = dict modèle.

    Gère les lignes de continuation : quand une ligne sans N° porte uniquement
    des actes médicaux (coupure de page), ceux-ci sont rattachés à la ligne
    précédente.
    """
    with pdfplumber.open(path) as pdf:
        facture = mois = None
        raw_rows = []
        for page in pdf.pages:
            txt = page.extract_text() or ""
            for line in txt.split("\n"):
                m = re.search(r"Facture N°\s*:\s*(\S+)", line)
                if m:
                    facture = m.group(1)
                m = re.search(r"Mois de prise en charge\s*:\s*(.+)", line)
                if m:
                    mois = unicodedata.normalize("NFC", m.group(1).strip())
            for table in page.extract_tables():
                for row in table:
                    c0 = str(row[0] or "").strip()
                    c4 = str(row[4] or "").strip()
                    if c0 == "N°":
                        continue  # en-tête de tableau
                    if c0 == "Total" or c4 == "Total":
                        continue  # ligne total (le libellé peut être en col 0 ou 4)
                    if not c0:
                        if c4 and raw_rows:
                            # continuation inter-page : rattacher les actes à la ligne précédente
                            prev = raw_rows[-1]
                            prev[4] = (prev[4] or "") + "\n" + c4
                        continue
                    raw_rows.append(list(row))

    lignes = []
    for row in raw_rows:
        num, date, matricule, nom_cell, actes_cell, montant, part, net = row[:8]

        # --- Nom / sous-société ---
        parts = [p.strip() for p in (nom_cell or "").split("\n") if p.strip()]
        sous = ""
        name_parts = []
        for p in parts:
            m = re.fullmatch(r"\((.+)\)", p)
            if m:
                sous = m.group(1).strip()
            else:
                name_parts.append(p)
        nom = " ".join(name_parts)
        nom_agent = f"{nom} ({sous})" if sous else nom

        # --- Actes médicaux ---
        actes = []
        total_actes = 0.0
        for a in (actes_cell or "").split("\n"):
            a = a.strip()
            if not a:
                continue
            m = re.match(r"^(.*?)\s*:\s*([\d\s\u00a0,\.]+)$", a)
            if m:
                lib, val = m.group(1).strip(), amount_to_float(m.group(2))
                total_actes += val
                actes.append(f"{lib} : {fmt_amount(val)}")
            else:
                actes.append(a)
        actes_str = "; ".join(actes)

        brut = amount_to_float(montant)
        tm = amount_to_float(part)
        net_pay = amount_to_float(net)

        # --- Société (préfixe du nom de fichier : BSA xxx.pdf -> BSA) ---
        societe = os.path.basename(path).split(" ")[0].upper()

        # --- Observations ---
        obs = f"Facture mensuelle soins ambulatoires - {mois}"
        if abs(total_actes - brut) > 0.01:
            obs += f" ; ATTENTION : actes détaillés ({fmt_amount(total_actes)} Ar) < montant facturé (écart {fmt_amount(brut - total_actes)} Ar)"

        lignes.append({
            "Numero_Facture": facture,
            "Date_Soins": parse_date(date),
            "Nom_Agent": nom_agent,
            "Matricule": str(matricule or "").strip(),
            "Societe": societe,
            "Sous_Societe": sous,
            "Acte_Medicale_Prix": actes_str,
            "Montant_Total_Brut": int(brut) if brut == int(brut) else brut,
            "Ticket_Moderateur": int(tm) if tm == int(tm) else tm,
            "Prise_En_Charge_Net": int(net_pay) if net_pay == int(net_pay) else net_pay,
            "Observations": obs,
        })
    return facture, mois, lignes


def style_sheet(ws):
    """Applique la mise en forme du modèle (largeurs, police, formats)."""
    model_ws = load_workbook(MODEL)["Modele_Prestations"]
    widths = {k: v.width for k, v in model_ws.column_dimensions.items()}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(name="Calibri", size=12)
    for r in range(2, ws.max_row + 1):
        for c in range(1, 12):
            ws.cell(row=r, column=c).font = openpyxl.styles.Font(name="Calibri", size=12)
        for c in (8, 9, 10):  # montants
            ws.cell(row=r, column=c).number_format = "#,##0"
    ws.freeze_panes = "A2"


def write_workbook(path, lignes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Modele_Prestations"
    ws.append(HEADERS)
    for l in lignes:
        ws.append([l[h] for h in HEADERS])
    style_sheet(ws)
    wb.save(path)


def month_of(path):
    """Nom du mois en majuscules à partir du nom de fichier '<CLIENT> <MOIS>.pdf'."""
    return os.path.basename(path).rsplit(".", 1)[0].split(" ", 1)[1].upper()


def process_client(societe, force=False):
    """Convertit tous les PDF d'un client et produit le fichier consolidé."""
    pattern = os.path.join(PDF_DIR, f"{societe} *.pdf")
    pdfs = [p for p in glob.glob(pattern) if month_of(p) in MONTH_ORDER]
    pdfs.sort(key=lambda p: MONTH_ORDER.index(month_of(p)))
    if not pdfs:
        print(f"!! Aucun PDF trouvé pour {societe} ({pattern})")
        return

    toutes = []
    for pdf_path in pdfs:
        facture, mois, lignes = parse_pdf(pdf_path)
        if not lignes:
            print(f"!! {os.path.basename(pdf_path)} : aucune ligne trouvée")
            continue
        stem = os.path.splitext(os.path.basename(pdf_path))[0]
        out = os.path.join(OUT_DIR, f"{stem}.xlsx")
        if os.path.exists(out) and not force:
            print(f"-- {stem}.xlsx : déjà existant, non écrasé (utilisez --force pour régénérer)")
        else:
            write_workbook(out, lignes)
            s_brut = sum(l["Montant_Total_Brut"] for l in lignes)
            s_tm = sum(l["Ticket_Moderateur"] for l in lignes)
            s_net = sum(l["Prise_En_Charge_Net"] for l in lignes)
            print(f"OK {stem}.xlsx : {facture} | {mois} | {len(lignes)} lignes | "
                  f"Brut {s_brut:,} | TM {s_tm:,} | Net {s_net:,}".replace(",", " "))
        # pour le consolidé, relire le fichier réellement en place (garde les modif. manuelles)
        if os.path.exists(out):
            ws = load_workbook(out)["Modele_Prestations"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                toutes.append(dict(zip(HEADERS, row)))

    if not toutes:
        print(f"!! {societe} : rien à consolider")
        return
    mois_premier = month_of(pdfs[0]).capitalize()
    mois_dernier = month_of(pdfs[-1]).capitalize()
    consolide = os.path.join(OUT_DIR, f"{societe} 2026 CONSOLIDE ({mois_premier}-{mois_dernier}).xlsx")
    if os.path.exists(consolide) and not force:
        print(f"-- {os.path.basename(consolide)} : déjà existant, non écrasé")
    else:
        write_workbook(consolide, toutes)
        print(f"OK {os.path.basename(consolide)} : {len(toutes)} lignes au total")


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    args = [a for a in sys.argv[1:] if a != "--force"]
    force = "--force" in sys.argv[1:]
    if args:
        clients = [a.upper() for a in args]
    else:  # découverte automatique des clients (préfixes des PDF)
        clients = sorted({os.path.basename(p).split(" ")[0].upper()
                          for p in glob.glob(os.path.join(PDF_DIR, "*.pdf"))})
    for societe in clients:
        process_client(societe, force)


if __name__ == "__main__":
    main()
