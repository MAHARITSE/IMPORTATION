# -*- coding: utf-8 -*-
"""
BSA_paiement_to_excel.py — société BSA
=======================================
Conversion des relevés de remboursements BSA (PDF) en fichiers Excel
selon le modèle "Modele_Import_Reglements_Decompte_Assurance.xlsx"
(feuille Modele_Reglements, 13 colonnes).

Format traité (propre à BSA) : RELEVE DE REMBOURSEMENTS DES FRAIS DE SANTE
  - ordre de virement en 1re page : N°, Lot, "A , le 17/04/2026",
    "virement de 928 750,00 MGA"
  - puis un bloc par remboursement :
      "1071921-1 ADHESION: 950179 RAKOTOARINAIVO CLOTAIRE CG Client: ..."
      "05/02/2026 RAKOTOARINAIVO ASSOCIATION DISPENSAIRE LOTERANA
       15 000,00 0,00 95,00 14 250,00 750,00 0,00"
      (date | nom patient | executant | FR.REELS | 1ERE MUT | Tx (%) |
       REMB | NON REMB | TPG*)
      Règle de cohérence BSA : le reste (non remboursé) = FR.REELS - REMB.
      Si FR.REELS = REMB (tout est remboursé) alors le reste = 0.
      Dans le PDF, NON_REMB inclut le TPG (ticket modérateur) :
      NON_REMB - TPG doit être égal au reste. Chaque ligne est contrôlée ;
      tout écart d'au moins 1 Ar est signalé (INCOHÉRENCE) et noté dans
      la colonne Motif_Observation.
      Montant_Exclu_Rejet = le reste (= 0 si FR.REELS = REMB).
      Ticket_Moderateur = Tx (%) (taux de prise en charge).
  - page finale : facture SALFA (n° FA-...) et "Total général".

Utilisation (double-clic sur BSA_paiement.bat, ou invite de commandes) :
    python BSA_paiement_to_excel.py                    # tous les PDF du sous-dossier PDF/
    python BSA_paiement_to_excel.py --force            # régénérer (écrase l'Excel existant)
    python BSA_paiement_to_excel.py "mon_releve.pdf"   # un seul PDF (nom ou chemin)

Sortie : BSA/<ANNEE>/<DATE_PAIEMENT> BSA <ANNEE> <PERIODE> MONTANT <MONTANT>Ar.xlsx
    (sous-dossier <ANNEE> = année du virement, créé automatiquement)
    exemple : BSA/2026/17-04-26 BSA 2026 27-01-26 à 23-02-26 MONTANT 928 750Ar.xlsx
    - DATE_PAIEMENT : date du virement (ligne "A , le 17/04/2026"), au format JJ-MM-AA
    - SOCIETE       : BSA (fixée dans ce script)
    - ANNEE         : année du virement (sert aussi de nom au sous-dossier)
    - PERIODE       : 1re et dernière date de soins du relevé, au format JJ-MM-AA
                      (une seule date si toutes les lignes sont du même jour)
    - MONTANT       : total payé (somme des REMB = montant du virement),
                      précédé du mot "MONTANT" et suivi de "Ar"

Numéro de facture SALFA (colonne Numero_Facture_Prescription) :
  - ancien format : un n° unique "FA-02/BFV/26-022" pour tout le relevé ;
  - format 2025   : un n° par décompte, lu en fin de chaque décompte :
        "Total décompte : 1015497"
        "Date facture: ... FACTURE SALFA TOLIARA N°006-25/BFV/BSA/SA"
    Chaque ligne Excel reçoit le n° de facture de SON décompte.

PDF en erreur : si un PDF ne peut pas être converti (PDF illisible, format
non reconnu, aucune ligne trouvée, date introuvable, erreur pendant la
création de l'Excel), il est DÉPLACÉ dans le sous-dossier BSA/ERREUR/
(créé automatiquement). Les PDF du dossier ERREUR ne sont pas retraités :
remettez le PDF dans PDF/ après correction pour réessayer.

Les fichiers Excel déjà existants ne sont PAS écrasés (protection des
modifications manuelles), sauf avec l'option --force.
"""
import re
import sys
import glob
import os
import shutil
import unicodedata
import pdfplumber
import openpyxl.styles
from openpyxl import Workbook, load_workbook

# Le script se trouve dans le dossier de la société BSA.
# Les PDF à convertir sont déposés dans le sous-dossier "PDF" de ce dossier
# (BSA/PDF/) ; à défaut, ils sont cherchés directement dans le dossier BSA.
# Le modèle Excel est dans le dossier parent "PAIEMENT CLIENT".
# Les Excel produits sont classés dans un sous-dossier au nom de l'année
# du règlement (ex : BSA/2026/...), créé automatiquement.
PDF_DIR = os.path.dirname(os.path.abspath(__file__))
PDF_SUBDIR = os.path.join(PDF_DIR, "PDF")
MODEL = os.path.join(os.path.dirname(PDF_DIR),
                     "Modele_Import_Reglements_Decompte_Assurance.xlsx")
SHEET = "Modele_Reglements"
SOCIETE = "BSA"

HEADERS = ["Ref_Decompte", "Date_Reglement", "Date_Soins", "Nom_Agent", "Matricule",
           "Numero_Facture_Prescription", "Code_Acte", "Libelle_Acte",
           "Montant_Reclame_Brut", "Ticket_Moderateur", "Montant_Paye_Regle",
           "Montant_Exclu_Rejet", "Motif_Observation"]

# Montant malgache : "15 000,00" / "0,00" / "112 500,00"
AMT = r"(?:\d{1,3}(?:[ \u00a0]\d{3})*|\d+),\d{2}"

# Ligne de données BSA : date + (nom + executant) + 6 montants
DATA_RE = re.compile(
    r"^(\d{2}/\d{2}/\d{4})\s+(.*?)\s+"
    + r"\s+".join(["(" + AMT + ")"] * 6) + r"$")

# Ligne d'en-tête de bloc BSA : "1071921-1 ADHESION: 950179 NOM... CG Client: ..."
BLOCK_RE = re.compile(r"^(\d{4,}-\d+)\s+ADHESION:\s*(\d+)\s+(.+)$")

# N° de facture SALFA : "FA-02/BFV/26-022" (ancien format) ou
# "N°006-25/BFV/BSA/SA" (format 2025). La ligne "Facture N°: N°006- ..." du
# PDF est tronquée (les montants suivent tout de suite) : seul le n° COMPLET
# de la ligne "Date facture: ..." est retenu.
FACTURE_RE = re.compile(r"(FA-\d{2}[-/][\w/\-]*\d|N°\s*\d{2,4}[-/][\w/\-]*\w)")

# Tokens qui marquent la fin d'un bloc (en-têtes de page, totaux, pied de page)
STOP_TOKENS = {"SALFATU", "TOLIARY", "MADAGASCAR", "Andraharo", "RELEVE",
               "Lot", "N°", "Banque", "Ville", "DATE", "AYANT-DROIT",
               "EXECUTANT", "FR.REELS", "TPG*", "REMBOURSEMENTS"}


def sans_accent(s):
    """'Février' -> 'FEVRIER' (majuscules sans accent, pour les comparaisons)."""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper()


def amount_to_float(s):
    """Montant d'un PDF -> nombre. '' / None / texte non numérique -> 0.0.

    Gère les deux notations rencontrées dans les PDF :
        française : '15 000,00' / '928 750'  -> 15000.0 / 928750.0
        anglaise  : '75,000' / '1,234.56'    -> 75000.0 / 1234.56
    """
    s = re.sub(r"[^\d.,-]", "", str(s or "").replace("\u00a0", " "))
    if not s:
        return 0.0
    virgule, point = s.rfind(","), s.rfind(".")
    dernier = max(virgule, point)
    if dernier >= 0 and s.count(",") + s.count(".") == 1 \
            and len(s) - dernier - 1 == 3:
        # un seul séparateur suivi de 3 chiffres = séparateur de milliers
        s = s.replace(",", "").replace(".", "")
    elif virgule > point:
        s = s.replace(".", "").replace(",", ".")   # virgule = décimales
    else:
        s = s.replace(",", "")                     # point = décimales
    try:
        return float(s)
    except ValueError:
        return 0.0


def fmt_amount(n):
    """928750.0 -> '928 750' ; 1234.5 -> '1 234,5'"""
    if abs(n - round(n)) < 0.005:
        return f"{int(round(n)):,}".replace(",", " ")
    return f"{n:,.1f}".replace(",", " ").replace(".", ",")


def num(s):
    """'15 000,00' -> 15000 (int si entier)"""
    v = amount_to_float(s)
    return int(v) if v == int(v) else v


def parse_date(d):
    """'02/01/2026' -> '2026-01-02'"""
    dd, mm, yy = d.strip().split("/")
    year = int(yy)
    year += 2000 if year < 100 else 0
    return f"{year:04d}-{mm}-{dd}"


# --------------------------------------------------------------------------
# Nom du fichier Excel de sortie
#   <DATE_PAIEMENT> <SOCIETE> <ANNEE> <PERIODE> MONTANT <MONTANT>Ar.xlsx
#   exemple : 17-04-26 BSA 2026 27-01-26 à 23-02-26 MONTANT 928 750Ar.xlsx
# Classé dans un sous-dossier au nom de l'année du règlement :
#   BSA/2026/17-04-26 BSA 2026 27-01-26 à 23-02-26 MONTANT 928 750Ar.xlsx
# --------------------------------------------------------------------------
# Caractères interdits dans un nom de fichier Windows
INVALIDES = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

DATE_ISO = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def date_courte(iso):
    """'2026-01-27' -> '27-01-26' (JJ-MM-AA)."""
    annee, mois, jour = iso.split("-")
    return f"{jour}-{mois}-{annee[2:]}"


def periode_soins(lignes, defaut=None):
    """Période couverte par le paiement : '27-01-26 à 23-02-26'.

    Prend la 1re et la dernière date de soins (colonne Date_Soins) des lignes
    du fichier. Une seule date si toutes les lignes tombent le même jour.
    À défaut (aucune date de soins lisible), la date de règlement.
    """
    dates = sorted(l.get("Date_Soins") or "" for l in lignes)
    dates = [d for d in dates if DATE_ISO.match(d)]
    if not dates:
        if not (defaut and DATE_ISO.match(defaut)):
            return "SANS DATE"
        dates = [defaut]
    debut, fin = dates[0], dates[-1]
    return date_courte(debut) if debut == fin \
        else f"{date_courte(debut)} à {date_courte(fin)}"


def nom_sortie(societe, date_reglement, lignes, montant):
    """Construit le nom du fichier Excel :

        <DATE_PAIEMENT> <SOCIETE> <ANNEE> <PERIODE> MONTANT <MONTANT>Ar.xlsx
        17-04-26 BSA 2026 27-01-26 à 23-02-26 MONTANT 928 750Ar.xlsx

    date_reglement : 'AAAA-MM-JJ' (date du virement / date comptable).
    """
    annee = date_reglement.split("-")[0]
    nom = (f"{date_courte(date_reglement)} {societe} {annee} "
           f"{periode_soins(lignes, date_reglement)} "
           f"MONTANT {fmt_amount(montant)}Ar")
    nom = re.sub(r"\s+", " ", INVALIDES.sub(" ", nom)).strip()
    return nom + ".xlsx"


def dossier_annee(date_reglement):
    """Chemin complet du dossier de l'année du règlement (créé si absent).

    Les Excel sont classés par année du règlement :
        BSA/2026/17-04-26 BSA 2026 27-01-26 à 23-02-26 MONTANT 928 750Ar.xlsx

    date_reglement : 'AAAA-MM-JJ'.
    """
    annee = date_reglement.split("-")[0]
    d = os.path.join(PDF_DIR, annee)
    os.makedirs(d, exist_ok=True)
    return d


def group_words(words, tol=2.5):
    """Regroupe les mots d'une page en lignes (par coordonnée verticale).

    Retourne une liste de dictionnaires : {text, x0, words} où words est
    [(x0, x1, texte), ...] trié de gauche à droite.
    """
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    groups = []
    cur, cur_top = [], None
    for w in words:
        if cur and w["top"] - cur_top > tol:
            groups.append(cur)
            cur, cur_top = [], None
        if not cur:
            cur_top = w["top"]
        cur.append(w)
    if cur:
        groups.append(cur)
    out = []
    for g in groups:
        g = sorted(g, key=lambda w: w["x0"])
        out.append({
            "text": " ".join(w["text"] for w in g),
            "x0": g[0]["x0"],
            "words": [(w["x0"], w["x1"], w["text"]) for w in g],
        })
    return out


def is_stop(t):
    """Vrai si la ligne est un en-tête de page / total / pied de page."""
    if re.fullmatch(r"\d+\s*/\s*\d+", t):          # pied de page "2 /4"
        return True
    if re.match(r"^(Facture|Total|Date facture|Nbre)\b", t):
        return True
    if re.fullmatch(AMT + r"(?:\s+" + AMT + r")*", t):  # ligne de montants seule
        return True
    return bool(set(t.split()) & STOP_TOKENS)


def full_pdf_text(pdf):
    return "\n".join((page.extract_text() or "") for page in pdf.pages)


def controle_reste(fr, remb, nonremb, tpg):
    """Contrôle de cohérence des montants d'une ligne BSA.

    Règle BSA : le reste (montant non remboursé, écrit dans
    Montant_Exclu_Rejet) = FR.REELS - REMB.
    Si FR.REELS = REMB (tout est remboursé) alors le reste = 0.

    Contrôle croisé avec les colonnes du PDF : dans le PDF BSA,
    NON_REMB inclut le TPG (ticket modérateur), donc NON_REMB - TPG
    doit être égal au reste. Tout écart d'au moins 1 Ar (ou un REMB
    supérieur à FR.REELS) est une incohérence.

    Retourne (reste, problemes) :
      - reste     : le reste à écrire dans Montant_Exclu_Rejet (>= 0) ;
      - problemes : liste de messages d'explication (vide si cohérent).
    """
    reste = fr - remb
    exclu_pdf = nonremb - tpg
    problemes = []
    if remb > fr + 0.5:
        problemes.append(f"REMB ({fmt_amount(remb)}) > FR.REELS ({fmt_amount(fr)})")
    if abs(reste - exclu_pdf) >= 1:
        problemes.append(
            f"le reste FR.REELS-REMB ({fmt_amount(reste)}) ≠ "
            f"NON_REMB-TPG ({fmt_amount(exclu_pdf)})")
    return max(0.0, reste), problemes


# --------------------------------------------------------------------------
# Format BSA : RELEVE DE REMBOURSEMENTS DES FRAIS DE SANTE
# --------------------------------------------------------------------------
def parse_bsa(pdf, nom_pdf):
    """Extrait le méta du relevé et les lignes de remboursements.

    Chaque ligne du PDF est un bloc :
      ligne 1 : "1071921-1 ADHESION: 950179 RAKOTOARINAIVO CLOTAIRE CG Client: ..."
      ligne 2 : "05/02/2026 RAKOTOARINAIVO ASSOCIATION DISPENSAIRE LOTERANA
                 15 000,00 0,00 95,00 14 250,00 750,00 0,00"
                 (date | nom patient | executant | FR.REELS | 1ERE MUT | Tx |
                  REMB | NON REMB | TPG*)
      Règle de cohérence : le reste = FR.REELS - REMB ; si FR.REELS = REMB
      alors le reste = 0. Le reste est contrôlé contre NON_REMB - TPG
      (le NON_REMB du PDF inclut le ticket modérateur) : tout écart
      d'au moins 1 Ar est signalé comme INCOHÉRENCE.
      Montant_Exclu_Rejet = le reste (= 0 si FR.REELS = REMB).
      Ticket_Moderateur = Tx (%) (taux de prise en charge BSA).
      lignes 3+ : suite du nom (colonne de gauche, x<125) puis libellé de
                 l'acte / médicament (x>=125), tant que la ligne n'est pas un
                 en-tête de page, un total ou le pied de page.
    """
    lines = []
    for page in pdf.pages:
        lines.extend(group_words(page.extract_words()))
    text = "\n".join(l["text"] for l in lines)

    # --- Métadonnées du relevé ---
    meta = {"ref": None, "lot": None, "date_reglement": None, "virement": None,
            "facture": None, "factures_decompte": {}, "nb_declare": None}
    m = re.search(r"N°\s*:\s*(\d+)", text)
    if m:
        meta["ref"] = m.group(1)
    m = re.search(r"Lot\s*:\s*(\d+)", text)
    if m:
        meta["lot"] = m.group(1)
    m = re.search(r"\ble\s+(\d{2}/\d{2}/\d{4})", text)   # "A , le 17/04/2026"
    if m:
        meta["date_reglement"] = m.group(1)
    m = re.search(r"virement de ([\d\u00a0 ]+),(\d{2})\s*MGA", text)
    if m:
        meta["virement"] = amount_to_float(m.group(1) + "," + m.group(2))
    m = re.search(r"Total général\s*:\s*(\d+)\s+(\d+)\s+", text)
    if m:
        meta["nb_declare"] = int(m.group(2))

    # --- N° de facture SALFA, par décompte ---
    #   "Total décompte : 1015497"
    #   "Date facture: ... FACTURE SALFA TOLIARA N°006-25/BFV/BSA/SA"
    # (dans cet ordre) -> factures_decompte["1015497"] = "N°006-25/BFV/BSA/SA".
    # Un n° lu avant tout "Total décompte" (ancien format : un seul pour le
    # relevé) est mémorisé dans meta["facture"] et servi à toutes les lignes.
    cur_dec = None
    for l in lines:
        t = l["text"]
        m = re.search(r"Total d[ée]compte\s*:?\s*(\d{4,})", t)
        if m:
            cur_dec = m.group(1)
            continue
        if "facture" not in t.lower():
            continue
        m = FACTURE_RE.search(t)
        if m:
            if cur_dec is not None:
                meta["factures_decompte"].setdefault(cur_dec, m.group(1))
            elif meta["facture"] is None:
                meta["facture"] = m.group(1)
    if meta["facture"] is None:      # ancien format : n° "FA-..." unique
        for cand in re.findall(r"n°\s*:?\s*(FA-[\w/\-]+)", text):
            if cand[-1].isdigit():   # le n° complet finit par un chiffre
                meta["facture"] = cand

    # --- Blocs de remboursements ---
    blocks = []
    cur = None
    for l in lines:
        t = l["text"]
        m = BLOCK_RE.match(t)
        if m:
            if cur:
                blocks.append(cur)
            # nom (x<255) / acte (x>=255) dans la ligne d'en-tête ;
            # le 1er numéro après "ADHESION:" est le matricule (déjà lu).
            names, actes = [], []
            state = "skip"  # skip -> nom/acte après le matricule
            for x0, x1, w in l["words"]:
                if w == "Client:":
                    break
                if state == "skip":
                    if w.isdigit() and int(x0) > 90:
                        state = "ok"
                    continue
                if state == "ok":
                    (actes if x0 >= 255 else names).append(w)
            cur = {"num": m.group(1), "matricule": m.group(2),
                   "nom_entete": " ".join(names), "acte": " ".join(actes),
                   "data": None, "nom_extra": [], "libelle": []}
            continue
        if cur is None:
            continue
        dm = DATA_RE.match(t)
        if dm and cur["data"] is None:
            cur["data"] = dm
            continue
        if is_stop(t):
            blocks.append(cur)
            cur = None
            continue
        if cur["data"] is None:
            continue  # bruit avant la ligne de données (en-têtes de colonne...)
        if l["x0"] < 125:
            cur["nom_extra"].append(t)   # suite du nom (colonne AYANT-DROIT)
        else:
            cur["libelle"].append(t)     # libellé de l'acte / médicament
    if cur:
        blocks.append(cur)

    # --- Construction des lignes Excel ---
    lignes = []
    nb_incoherences = 0     # lignes dont les montants ne respectent pas la règle
    for b in blocks:
        if b["data"] is None:
            print(f"!! {nom_pdf} : bloc {b['num']} sans ligne de données -> ignoré")
            continue
        dm = b["data"]
        date, milieu, fr, mut, tx, remb, nonremb, tpg = dm.groups()

        # Nom du patient : 1re ligne sur la ligne de données (avant "ASSOCIATION")
        # + lignes de suite dans la colonne du nom.
        nom = re.split(r"\s+ASSOCIATION\b", milieu, 1)[0].strip()
        if b["nom_extra"]:
            nom = (nom + " " + " ".join(b["nom_extra"])).strip()
        if not nom:
            nom = b["nom_entete"]

        tx_v = amount_to_float(tx)
        motif = f"Prise en charge : {tx_v:g}%"
        if amount_to_float(mut) > 0:
            motif += f" ; 1ère mutuelle : {fmt_amount(amount_to_float(mut))} Ar"

        # --- Cohérence des montants (règle BSA) -------------------------
        # Le reste (non remboursé) = FR.REELS - REMB.
        # Si FR.REELS = REMB alors le reste = 0.
        # Contrôle croisé : NON_REMB - TPG doit donner le même reste.
        fr_v = amount_to_float(fr)
        remb_v = amount_to_float(remb)
        nonremb_v = amount_to_float(nonremb)
        tpg_v = amount_to_float(tpg)
        exclu_v, problemes = controle_reste(fr_v, remb_v, nonremb_v, tpg_v)
        if problemes:
            nb_incoherences += 1
            for p in problemes:
                print(f"   !! INCOHÉRENCE {nom_pdf} : bloc {b['num']} "
                      f"{date} {nom} : {p} "
                      f"(FR.REELS={fmt_amount(fr_v)}, "
                      f"REMB={fmt_amount(remb_v)}, "
                      f"NON_REMB={fmt_amount(nonremb_v)}, "
                      f"TPG={fmt_amount(tpg_v)})")
            motif += " !! INCOHÉRENCE : " + " ; ".join(problemes)

        # Facture SALFA de CE décompte (format 2025), sinon celle du relevé.
        decompte = b["num"].split("-")[0]
        facture = (meta["factures_decompte"].get(decompte)
                   or meta["facture"] or "")

        lignes.append({
            "Ref_Decompte": meta["ref"] or "",
            "Date_Reglement": parse_date(meta["date_reglement"]) if meta["date_reglement"] else "",
            "Date_Soins": parse_date(date),
            "Nom_Agent": nom,
            "Matricule": b["matricule"],
            "Numero_Facture_Prescription": facture,
            "Code_Acte": b["acte"],
            "Libelle_Acte": " ".join(b["libelle"]),
            "Montant_Reclame_Brut": num(fr),
            "Ticket_Moderateur": num(tx),
            "Montant_Paye_Regle": num(remb),
            "Montant_Exclu_Rejet": int(exclu_v) if exclu_v == int(exclu_v)
                                    else exclu_v,   # le reste (= 0 si FR.REELS = REMB)
            "Motif_Observation": motif,
        })
    meta["nb_incoherences"] = nb_incoherences
    return meta, lignes


# --------------------------------------------------------------------------
# Écriture Excel (mise en forme du modèle)
# --------------------------------------------------------------------------
def style_sheet(ws):
    model_ws = load_workbook(MODEL)[SHEET]
    widths = {k: v.width for k, v in model_ws.column_dimensions.items()}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(1, ws.max_row + 1):
        for c in range(1, 14):
            ws.cell(row=r, column=c).font = openpyxl.styles.Font(name="Calibri", size=12)
    for r in range(2, ws.max_row + 1):
        for c in (9, 10, 11, 12):  # montants
            ws.cell(row=r, column=c).number_format = "#,##0"
    ws.freeze_panes = "A2"


def write_workbook(path, lignes):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(HEADERS)
    for l in lignes:
        ws.append([l[h] for h in HEADERS])
    style_sheet(ws)
    wb.save(path)


# --------------------------------------------------------------------------
# Dossier des PDF en échec : BSA/ERREUR (créé automatiquement si besoin).
# Tout PDF impossible à convertir y est DÉPLACÉ, pour ne pas le confondre
# avec les PDF en attente de conversion. Les PDF du dossier ERREUR ne sont
# jamais retraités par les conversions suivantes : après correction, on
# remet le PDF dans PDF/ pour réessayer.
# --------------------------------------------------------------------------
ERREUR_DIRNAME = "ERREUR"
ERREURS = []        # [(nom du PDF, raison)] — récapitulatif en fin de traitement


def dossier_erreur():
    """Chemin du sous-dossier BSA/ERREUR (créé si absent)."""
    d = os.path.join(PDF_DIR, ERREUR_DIRNAME)
    os.makedirs(d, exist_ok=True)
    return d


def est_dans_erreur(pdf_path):
    """Vrai si le PDF se trouve déjà dans le sous-dossier ERREUR."""
    parties = os.path.normpath(os.path.abspath(pdf_path)).split(os.sep)
    return ERREUR_DIRNAME in parties


def deplacer_pdf_en_erreur(pdf_path, raison):
    """Déplace un PDF impossible à convertir dans BSA/ERREUR.

    Ne bouge pas le PDF si le déplacement échoue (fichier verrouillé...) :
    le problème est affiché et le PDF sera retraité au prochain lancement.
    """
    dest = os.path.join(dossier_erreur(), os.path.basename(pdf_path))
    base, ext = os.path.splitext(dest)
    n = 1
    while os.path.exists(dest):     # ne pas écraser un PDF déjà signalé
        dest = f"{base} ({n}){ext}"
        n += 1
    try:
        shutil.move(pdf_path, dest)
        print(f"   -> PDF déplacé dans {ERREUR_DIRNAME} : "
              f"{os.path.relpath(dest, PDF_DIR)}")
        ERREURS.append((os.path.basename(pdf_path), raison))
    except Exception as e:
        print(f"   !! déplacement impossible vers {ERREUR_DIRNAME} : {e}")
        ERREURS.append((os.path.basename(pdf_path), raison + f" (non déplacé : {e})"))


def recap_erreurs():
    """Affiche le récapitulatif des PDF en erreur (fin de traitement)."""
    if ERREURS:
        print(f"\n== {len(ERREURS)} PDF en erreur, déplacés dans "
              f"le sous-dossier {ERREUR_DIRNAME} de {SOCIETE} ==")
        for nom, raison in ERREURS:
            print(f"   - {nom} : {raison}")


# --------------------------------------------------------------------------
def main():
    args = [a for a in sys.argv[1:] if a != "--force"]
    force = "--force" in sys.argv[1:]

    # PDF à convertir : ceux passés en argument, sinon tous les PDF du dossier.
    pdfs = []
    if args:
        for a in args:
            if os.path.isabs(a):
                p = a
            else:
                # nom relatif : d'abord dans le sous-dossier PDF/, sinon
                # directement dans le dossier de la société
                p = os.path.join(PDF_SUBDIR, a)
                if not os.path.isfile(p):
                    p = os.path.join(PDF_DIR, a)
            if os.path.isfile(p) and p.lower().endswith(".pdf"):
                pdfs.append(os.path.abspath(p))
            else:
                print(f"!! PDF introuvable : {a}")
    else:
        # PDF déposés dans le sous-dossier PDF/ ; à défaut, directement dans
        # le dossier de la société, puis recherche récursive.
        pdfs = sorted(glob.glob(os.path.join(PDF_SUBDIR, "*.pdf")))
        if not pdfs:
            pdfs = sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf")))
        if not pdfs:
            pdfs = sorted(glob.glob(os.path.join(PDF_DIR, "**", "*.pdf"), recursive=True))
    if not pdfs:
        if args:
            print("!! Aucun des PDF demandés n'a été trouvé")
        else:
            print(f"!! Aucun PDF trouvé dans {PDF_SUBDIR}")
        recap_erreurs()
        return
    # Ne pas retraiter les PDF déjà mis de côté dans ERREUR
    pdfs = [p for p in pdfs if not est_dans_erreur(p)]

    for pdf_path in pdfs:
        nom_pdf = os.path.basename(pdf_path)
        try:
            pdf = pdfplumber.open(pdf_path)
        except Exception as e:
            print(f"!! {nom_pdf} : PDF illisible ({e})")
            deplacer_pdf_en_erreur(pdf_path, "PDF illisible ou corrompu")
            continue
        try:
            with pdf:
                # Dossier BSA → parseur BSA. Le contenu du PDF ne change rien.
                meta, lignes = parse_bsa(pdf, nom_pdf)
        except Exception as e:
            print(f"!! {nom_pdf} : erreur pendant la lecture du PDF "
                  f"({type(e).__name__}: {e})")
            deplacer_pdf_en_erreur(pdf_path, "erreur de lecture (format non reconnu ?)")
            continue

        if not lignes:
            print(f"!! {nom_pdf} : aucune ligne de règlement trouvée "
                  f"(format non reconnu ?)")
            deplacer_pdf_en_erreur(pdf_path, "aucune ligne trouvée dans le PDF")
            continue

        # --- Nom du fichier : DATE_PAIEMENT BSA ANNEE PERIODE MONTANT <montant>Ar,
        #     classé dans le sous-dossier de l'année du règlement ---
        dr = (meta.get("date_reglement") or "").strip()
        if not re.fullmatch(r"\d{2}/\d{2}/\d{4}", dr):
            print(f"!! {nom_pdf} : date de règlement introuvable dans le PDF")
            deplacer_pdf_en_erreur(pdf_path, "date de règlement introuvable")
            continue
        date_reglement = parse_date(dr)          # 'AAAA-MM-JJ'

        total_paye = sum(l["Montant_Paye_Regle"] for l in lignes)
        ref = f"relevé N° {meta['ref']}" if meta["ref"] else "relevé"
        if meta["virement"] is not None:
            ok = abs(total_paye - meta["virement"]) < 1
            if ok:
                print(f"   contrôle : {fmt_amount(total_paye)} Ar payés "
                      f"= montant du virement ({fmt_amount(meta['virement'])} Ar)  OK")
            else:
                print(f"   !! ATTENTION : {fmt_amount(total_paye)} Ar payés "
                      f"≠ montant du virement ({fmt_amount(meta['virement'])} Ar)")
        if meta["nb_declare"] is not None and meta["nb_declare"] != len(lignes):
            print(f"   !! ATTENTION : {len(lignes)} lignes lues, "
                  f"{meta['nb_declare']} déclarées dans le total général du relevé")

        # --- Contrôle de cohérence des montants -------------------------
        # Règle BSA : le reste = FR.REELS - REMB ; si FR.REELS = REMB
        # alors le reste doit être 0 (contrôlé aussi contre NON_REMB - TPG).
        nb_incoh = meta.get("nb_incoherences", 0)
        if nb_incoh:
            print(f"   !! ATTENTION : {nb_incoh} ligne(s) incohérente(s) — "
                  f"règle : si FR.REELS = REMB alors le reste = 0 "
                  f"(détail ci-dessus, et noté dans Motif_Observation)")
        else:
            print(f"   cohérence : {len(lignes)} lignes vérifiées — "
                  f"si FR.REELS = REMB alors le reste = 0  OK")

        out = os.path.join(dossier_annee(date_reglement),
                           nom_sortie(SOCIETE, date_reglement, lignes, total_paye))
        relatif = os.path.relpath(out, PDF_DIR)
        if os.path.exists(out) and not force:
            print(f"-- {relatif} : existe déjà, non écrasé "
                  f"(--force pour régénérer)  [{nom_pdf}]")
            continue
        try:
            write_workbook(out, lignes)
        except Exception as e:
            print(f"!! {nom_pdf} : erreur pendant la création de l'Excel "
                  f"({type(e).__name__}: {e})")
            deplacer_pdf_en_erreur(pdf_path, "erreur de création de l'Excel")
            continue
        print(f"OK {relatif} : {ref} | {len(lignes)} lignes | "
              f"Payé {fmt_amount(total_paye)} Ar  <- {nom_pdf}")

    recap_erreurs()


if __name__ == "__main__":
    main()
