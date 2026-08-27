# -*- coding: utf-8 -*-
"""
MCI_CARE_paiement_to_excel.py — société MCI CARE
================================================
Conversion des décomptes de règlement MCI CARE (PDF) en fichiers Excel
selon le modèle "Modele_Import_Reglements_Decompte_Assurance.xlsx"
(feuille Modele_Reglements, 13 colonnes).

Format traité (propre à MCI CARE) : DECOMPTE DE REGLEMENT FACTURES
  - en-tête : Facture #, Garant, Date comptable, Total prestataire
  - tableau par bénéficiaire :
      Matricule | Bénéficiaire | Date de soins | Actes | Montant réclamé |
      (#)Mtt non remboursé | Base décomptée | Ticket Modérateur |
      Montant réglé

Ref_Decompte : le n° de facture du décompte, BRUT, comme les autres sociétés
  (aucun suffixe « /<n>L » avec le nombre de lignes) :
      FA-02/MCI/26-047  ->  Ref_Decompte = FA-02/MCI/26-047

Utilisation (double-clic sur MCI_CARE_paiement.bat, ou invite de commandes) :
    python MCI_CARE_paiement_to_excel.py                    # tous les PDF du sous-dossier PDF/
    python MCI_CARE_paiement_to_excel.py --force            # régénérer (écrase l'Excel existant)
    python MCI_CARE_paiement_to_excel.py "mon_decompte.pdf" # un seul PDF (nom ou chemin)

Sortie : MCI CARE/<ANNEE>/<DATE_PAIEMENT> MCI CARE <ANNEE> <PERIODE> MONTANT <MONTANT>Ar.xlsx
    (sous-dossier <ANNEE> = année de la date comptable, créé automatiquement)
    exemple : MCI CARE/2026/02-05-26 MCI CARE 2026 02-03-26 à 31-03-26 MONTANT 471 140Ar.xlsx
    - DATE_PAIEMENT : date comptable du décompte, au format JJ-MM-AA
    - SOCIETE       : MCI CARE (fixée dans ce script)
    - ANNEE         : année de la date comptable (sert aussi de nom au sous-dossier)
    - PERIODE       : 1re et dernière date de soins du décompte, au format JJ-MM-AA
                      (une seule date si toutes les lignes sont du même jour)
    - MONTANT       : total payé (Total prestataire), précédé du mot "MONTANT"
                      et suivi de "Ar"

PDF en erreur : si un PDF ne peut pas être converti (PDF illisible, format
non reconnu, aucune ligne trouvée, date introuvable, erreur pendant la
création de l'Excel), il est DÉPLACÉ dans le sous-dossier
MCI CARE/ERREUR/ (créé automatiquement). Les PDF du dossier ERREUR ne sont
pas retraités : remettez le PDF dans PDF/ après correction pour réessayer.

Les fichiers Excel déjà existants ne sont PAS écrasés (protection des
modifications manuelles), sauf avec l'option --force.
"""
import re
import sys
import glob
import os
import shutil
import pdfplumber
import openpyxl.styles
from openpyxl import Workbook, load_workbook

# Le script se trouve dans le dossier de la société MCI CARE.
# Les PDF à convertir sont déposés dans le sous-dossier "PDF" de ce dossier
# (MCI CARE/PDF/) ; à défaut, ils sont cherchés directement dans MCI CARE.
# Le modèle Excel est dans le dossier parent "PAIEMENT CLIENT".
# Les Excel produits sont classés dans un sous-dossier au nom de l'année
# de la date comptable (ex : MCI CARE/2026/...), créé automatiquement.
PDF_DIR = os.path.dirname(os.path.abspath(__file__))
PDF_SUBDIR = os.path.join(PDF_DIR, "PDF")
MODEL = os.path.join(os.path.dirname(PDF_DIR),
                     "Modele_Import_Reglements_Decompte_Assurance.xlsx")
SHEET = "Modele_Reglements"
SOCIETE = "MCI CARE"

HEADERS = ["Ref_Decompte", "Date_Reglement", "Date_Soins", "Nom_Agent", "Matricule",
           "Numero_Facture_Prescription", "Code_Acte", "Libelle_Acte",
           "Montant_Reclame_Brut", "Ticket_Moderateur", "Montant_Paye_Regle",
           "Montant_Exclu_Rejet", "Motif_Observation"]


def amount_to_float(s):
    """Montant d'un PDF -> nombre. '' / None / texte non numérique -> 0.0.

    Gère les deux notations rencontrées dans les PDF :
        française : '15 000,00' / '928 750'  -> 15000.0 / 928750.0
        anglaise  : '75,000' / '1,234.56'    -> 75000.0 / 1234.56
    """
    s = re.sub(r"[^\d.,-]", "", str(s or "").replace("\u00a0", " "))
    if not s:
        return 0.0
    virgule, point = s.rfind(","), s.rfind(".")
    dernier = max(virgule, point)
    if dernier >= 0 and s.count(",") + s.count(".") == 1 \
            and len(s) - dernier - 1 == 3:
        # un seul séparateur suivi de 3 chiffres = séparateur de milliers
        s = s.replace(",", "").replace(".", "")
    elif virgule > point:
        s = s.replace(".", "").replace(",", ".")   # virgule = décimales
    else:
        s = s.replace(",", "")                     # point = décimales
    try:
        return float(s)
    except ValueError:
        return 0.0


def fmt_amount(n):
    """471140.0 -> '471 140' ; 1234.5 -> '1 234,5'"""
    if abs(n - round(n)) < 0.005:
        return f"{int(round(n)):,}".replace(",", " ")
    return f"{n:,.1f}".replace(",", " ").replace(".", ",")


# Montant dans un décompte MCI : notation française ("1 153 900", "15 000,00")
# ou anglaise ("75,000", "1,234.56") selon l'édition du PDF.
MONTANT_RE = re.compile(
    r"\d{1,3}(?!\d)(?:[ \u00a0]\d{3})*(?:[.,]\d{3})?(?:[.,]\d{2})?"
    r"|\d+(?:[.,]\d+)?")


def dernier_montant(ligne):
    """Dernier montant d'une ligne de total = colonne "Montant réglé"."""
    trouves = MONTANT_RE.findall(ligne)
    return amount_to_float(trouves[-1]) if trouves else 0.0


def num(s):
    """'15 000,00' -> 15000 (int si entier)"""
    v = amount_to_float(s)
    return int(v) if v == int(v) else v


def parse_date(d):
    """'02/03/2026' -> '2026-03-02'"""
    dd, mm, yy = d.strip().split("/")
    year = int(yy)
    year += 2000 if year < 100 else 0
    return f"{year:04d}-{mm}-{dd}"


# --------------------------------------------------------------------------
# Nom du fichier Excel de sortie
#   <DATE_PAIEMENT> <SOCIETE> <ANNEE> <PERIODE> MONTANT <MONTANT>Ar.xlsx
#   exemple : 02-05-26 MCI CARE 2026 02-03-26 à 31-03-26 MONTANT 471 140Ar.xlsx
# Classé dans un sous-dossier au nom de l'année du règlement :
#   MCI CARE/2026/02-05-26 MCI CARE 2026 02-03-26 à 31-03-26 MONTANT 471 140Ar.xlsx
# --------------------------------------------------------------------------
# Caractères interdits dans un nom de fichier Windows
INVALIDES = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

DATE_ISO = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def date_courte(iso):
    """'2026-03-02' -> '02-03-26' (JJ-MM-AA)."""
    annee, mois, jour = iso.split("-")
    return f"{jour}-{mois}-{annee[2:]}"


def periode_soins(lignes, defaut=None):
    """Période couverte par le paiement : '02-03-26 à 31-03-26'.

    Prend la 1re et la dernière date de soins (colonne Date_Soins) des lignes
    du fichier. Une seule date si toutes les lignes tombent le même jour.
    À défaut (aucune date de soins lisible), la date de règlement.
    """
    dates = sorted(l.get("Date_Soins") or "" for l in lignes)
    dates = [d for d in dates if DATE_ISO.match(d)]
    if not dates:
        if not (defaut and DATE_ISO.match(defaut)):
            return "SANS DATE"
        dates = [defaut]
    debut, fin = dates[0], dates[-1]
    return date_courte(debut) if debut == fin \
        else f"{date_courte(debut)} à {date_courte(fin)}"


def nom_sortie(societe, date_reglement, lignes, montant):
    """Construit le nom du fichier Excel :

        <DATE_PAIEMENT> <SOCIETE> <ANNEE> <PERIODE> MONTANT <MONTANT>Ar.xlsx
        02-05-26 MCI CARE 2026 02-03-26 à 31-03-26 MONTANT 471 140Ar.xlsx

    date_reglement : 'AAAA-MM-JJ' (date comptable du décompte).
    """
    annee = date_reglement.split("-")[0]
    nom = (f"{date_courte(date_reglement)} {societe} {annee} "
           f"{periode_soins(lignes, date_reglement)} "
           f"MONTANT {fmt_amount(montant)}Ar")
    nom = re.sub(r"\s+", " ", INVALIDES.sub(" ", nom)).strip()
    return nom + ".xlsx"


def dossier_annee(date_reglement):
    """Chemin complet du dossier de l'année du règlement (créé si absent).

    Les Excel sont classés par année de la date comptable :
        MCI CARE/2026/02-05-26 MCI CARE 2026 02-03-26 à 31-03-26 MONTANT 471 140Ar.xlsx

    date_reglement : 'AAAA-MM-JJ'.
    """
    annee = date_reglement.split("-")[0]
    d = os.path.join(PDF_DIR, annee)
    os.makedirs(d, exist_ok=True)
    return d


def full_pdf_text(pdf):
    return "\n".join((page.extract_text() or "") for page in pdf.pages)


# --------------------------------------------------------------------------
# Format MCI : DECOMPTE DE REGLEMENT FACTURES
# --------------------------------------------------------------------------
def parse_mci(pdf, nom_pdf):
    """Extrait le méta du décompte et les lignes du tableau Matricule...Montant réglé."""
    text = full_pdf_text(pdf)

    meta = {"facture": None, "date_reglement": None, "garant": None,
            "total_prestataire": None}
    m = re.search(r"Facture #\s*:\s*(\S+)", text)
    if m:
        meta["facture"] = m.group(1)
    m = re.search(r"Date comptable:\s*(\d{2}/\d{2}/\d{4})", text)
    if not m:
        m = re.search(r"Edition\s*:\s*(\d{2}/\d{2}/\d{4})", text)
    if m:
        meta["date_reglement"] = m.group(1)
    m = re.search(r"Garant:\s*(\S+)\s+([^\n]+)", text)
    if m:
        meta["garant"] = m.group(1) + " " + m.group(2).strip()
    # "Total prestataire:" apparait une fois par page (une page = un
    # établissement : PHIE, LABO, IMAGERIE, DISPENSAIRE...). On additionne
    # tous ces totaux pour avoir le total payé du décompte complet.
    totaux = [dernier_montant(l)
              for l in re.findall(r"Total prestataire:([^\n]*)", text)]
    if totaux:
        meta["total_prestataire"] = sum(totaux)

    lignes = []
    for page in pdf.pages:
        for table in page.extract_tables():
            if not table or not str(table[0][0] or "").strip().startswith("Matricule"):
                continue
            for row in table[1:]:
                c = [str(x or "").strip() for x in row[:9]]
                # Ligne de données = matricule numérique + date de soins ;
                # sinon : ligne de total, "999 EXCLUSION...", description...
                if len(c) < 9 or not c[0].isdigit() \
                        or not re.fullmatch(r"\d{2}/\d{2}/\d{4}", c[2]):
                    continue
                (matricule, benef, date, acte, reclame,
                 nonremb, base, tm, regle) = c
                tm_montant = round(amount_to_float(reclame) - amount_to_float(regle))
                obs = f"Ticket modérateur {tm}"
                if amount_to_float(nonremb) > 0:
                    obs += f" ; Montant non remboursé : {fmt_amount(amount_to_float(nonremb))} Ar"
                lignes.append({
                    "Ref_Decompte": meta["facture"] or "",
                    "Date_Reglement": parse_date(meta["date_reglement"]) if meta["date_reglement"] else "",
                    "Date_Soins": parse_date(date),
                    "Nom_Agent": benef,
                    "Matricule": matricule,
                    "Numero_Facture_Prescription": meta["facture"] or "",
                    "Code_Acte": acte,
                    "Libelle_Acte": "",
                    "Montant_Reclame_Brut": num(reclame),
                    "Ticket_Moderateur": tm_montant,
                    "Montant_Paye_Regle": num(regle),
                    "Montant_Exclu_Rejet": num(nonremb),
                    "Motif_Observation": obs,
                })
    # Ref_Decompte = n° de facture du décompte, BRUT (comme BSA / ASCOMA) :
    # pas de suffixe « /<n>L » indiquant le nombre de lignes.
    return meta, lignes


# --------------------------------------------------------------------------
# Écriture Excel (mise en forme du modèle)
# --------------------------------------------------------------------------
def style_sheet(ws):
    model_ws = load_workbook(MODEL)[SHEET]
    widths = {k: v.width for k, v in model_ws.column_dimensions.items()}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(1, ws.max_row + 1):
        for c in range(1, 14):
            ws.cell(row=r, column=c).font = openpyxl.styles.Font(name="Calibri", size=12)
    for r in range(2, ws.max_row + 1):
        for c in (9, 10, 11, 12):  # montants
            ws.cell(row=r, column=c).number_format = "#,##0"
    ws.freeze_panes = "A2"


def write_workbook(path, lignes):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(HEADERS)
    for l in lignes:
        ws.append([l[h] for h in HEADERS])
    style_sheet(ws)
    wb.save(path)


# --------------------------------------------------------------------------
# Dossier des PDF en échec : MCI CARE/ERREUR (créé automatiquement si besoin).
# Tout PDF impossible à convertir y est DÉPLACÉ, pour ne pas le confondre
# avec les PDF en attente de conversion. Les PDF du dossier ERREUR ne sont
# jamais retraités par les conversions suivantes : après correction, on
# remet le PDF dans PDF/ pour réessayer.
# --------------------------------------------------------------------------
ERREUR_DIRNAME = "ERREUR"
ERREURS = []        # [(nom du PDF, raison)] — récapitulatif en fin de traitement


def dossier_erreur():
    """Chemin du sous-dossier MCI CARE/ERREUR (créé si absent)."""
    d = os.path.join(PDF_DIR, ERREUR_DIRNAME)
    os.makedirs(d, exist_ok=True)
    return d


def est_dans_erreur(pdf_path):
    """Vrai si le PDF se trouve déjà dans le sous-dossier ERREUR."""
    parties = os.path.normpath(os.path.abspath(pdf_path)).split(os.sep)
    return ERREUR_DIRNAME in parties


def deplacer_pdf_en_erreur(pdf_path, raison):
    """Déplace un PDF impossible à convertir dans MCI CARE/ERREUR.

    Ne bouge pas le PDF si le déplacement échoue (fichier verrouillé...) :
    le problème est affiché et le PDF sera retraité au prochain lancement.
    """
    dest = os.path.join(dossier_erreur(), os.path.basename(pdf_path))
    base, ext = os.path.splitext(dest)
    n = 1
    while os.path.exists(dest):     # ne pas écraser un PDF déjà signalé
        dest = f"{base} ({n}){ext}"
        n += 1
    try:
        shutil.move(pdf_path, dest)
        print(f"   -> PDF déplacé dans {ERREUR_DIRNAME} : "
              f"{os.path.relpath(dest, PDF_DIR)}")
        ERREURS.append((os.path.basename(pdf_path), raison))
    except Exception as e:
        print(f"   !! déplacement impossible vers {ERREUR_DIRNAME} : {e}")
        ERREURS.append((os.path.basename(pdf_path), raison + f" (non déplacé : {e})"))


def recap_erreurs():
    """Affiche le récapitulatif des PDF en erreur (fin de traitement)."""
    if ERREURS:
        print(f"\n== {len(ERREURS)} PDF en erreur, déplacés dans "
              f"le sous-dossier {ERREUR_DIRNAME} de {SOCIETE} ==")
        for nom, raison in ERREURS:
            print(f"   - {nom} : {raison}")


# --------------------------------------------------------------------------
def main():
    args = [a for a in sys.argv[1:] if a != "--force"]
    force = "--force" in sys.argv[1:]

    # PDF à convertir : ceux passés en argument, sinon tous les PDF du dossier.
    pdfs = []
    if args:
        for a in args:
            if os.path.isabs(a):
                p = a
            else:
                # nom relatif : d'abord dans le sous-dossier PDF/, sinon
                # directement dans le dossier de la société
                p = os.path.join(PDF_SUBDIR, a)
                if not os.path.isfile(p):
                    p = os.path.join(PDF_DIR, a)
            if os.path.isfile(p) and p.lower().endswith(".pdf"):
                pdfs.append(os.path.abspath(p))
            else:
                print(f"!! PDF introuvable : {a}")
    else:
        # PDF déposés dans le sous-dossier PDF/ ; à défaut, directement dans
        # le dossier de la société, puis recherche récursive.
        pdfs = sorted(glob.glob(os.path.join(PDF_SUBDIR, "*.pdf")))
        if not pdfs:
            pdfs = sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf")))
        if not pdfs:
            pdfs = sorted(glob.glob(os.path.join(PDF_DIR, "**", "*.pdf"), recursive=True))
    if not pdfs:
        if args:
            print("!! Aucun des PDF demandés n'a été trouvé")
        else:
            print(f"!! Aucun PDF trouvé dans {PDF_SUBDIR}")
        recap_erreurs()
        return
    # Ne pas retraiter les PDF déjà mis de côté dans ERREUR
    pdfs = [p for p in pdfs if not est_dans_erreur(p)]

    for pdf_path in pdfs:
        nom_pdf = os.path.basename(pdf_path)
        try:
            pdf = pdfplumber.open(pdf_path)
        except Exception as e:
            print(f"!! {nom_pdf} : PDF illisible ({e})")
            deplacer_pdf_en_erreur(pdf_path, "PDF illisible ou corrompu")
            continue
        try:
            with pdf:
                # Dossier MCI CARE → parseur MCI CARE. Le contenu du PDF ne change rien.
                meta, lignes = parse_mci(pdf, nom_pdf)
        except Exception as e:
            print(f"!! {nom_pdf} : erreur pendant la lecture du PDF "
                  f"({type(e).__name__}: {e})")
            deplacer_pdf_en_erreur(pdf_path, "erreur de lecture (format non reconnu ?)")
            continue

        if not lignes:
            print(f"!! {nom_pdf} : aucune ligne de règlement trouvée "
                  f"(format non reconnu ?)")
            deplacer_pdf_en_erreur(pdf_path, "aucune ligne trouvée dans le PDF")
            continue

        # --- Nom du fichier : DATE_PAIEMENT MCI CARE ANNEE PERIODE MONTANT <montant>Ar,
        #     classé dans le sous-dossier de l'année du règlement ---
        dr = (meta.get("date_reglement") or "").strip()
        if not re.fullmatch(r"\d{2}/\d{2}/\d{4}", dr):
            print(f"!! {nom_pdf} : date comptable introuvable dans le PDF")
            deplacer_pdf_en_erreur(pdf_path, "date comptable introuvable")
            continue
        date_reglement = parse_date(dr)          # 'AAAA-MM-JJ'

        # Montant payé = somme des "Montant réglé" des lignes (le total
        # prestataire du PDF sert seulement de contrôle, voir ci-dessous).
        total_paye = sum(l["Montant_Paye_Regle"] for l in lignes)
        ref = f"facture {meta['facture']}" if meta["facture"] else "décompte"
        if meta["total_prestataire"] is not None:
            somme = sum(l["Montant_Paye_Regle"] for l in lignes)
            if abs(somme - meta["total_prestataire"]) >= 1:
                print(f"   !! ATTENTION : {fmt_amount(somme)} Ar en lignes "
                      f"≠ total prestataire ({fmt_amount(meta['total_prestataire'])} Ar)")

        out = os.path.join(dossier_annee(date_reglement),
                           nom_sortie(SOCIETE, date_reglement, lignes, total_paye))
        relatif = os.path.relpath(out, PDF_DIR)
        if os.path.exists(out) and not force:
            print(f"-- {relatif} : existe déjà, non écrasé "
                  f"(--force pour régénérer)  [{nom_pdf}]")
            continue
        try:
            write_workbook(out, lignes)
        except Exception as e:
            print(f"!! {nom_pdf} : erreur pendant la création de l'Excel "
                  f"({type(e).__name__}: {e})")
            deplacer_pdf_en_erreur(pdf_path, "erreur de création de l'Excel")
            continue
        print(f"OK {relatif} : {ref} | {len(lignes)} lignes | "
              f"Payé {fmt_amount(total_paye)} Ar  <- {nom_pdf}")

    recap_erreurs()


if __name__ == "__main__":
    main()
