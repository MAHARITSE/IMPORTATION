# -*- coding: utf-8 -*-
"""
Conversion des factures PDF SALFA en fichiers Excel selon le modèle
"Modele_Import.xlsx" (feuille Modele_Prestations, 11 colonnes).

Usage :
    python pdf_to_excel.py            # convertit tous les PDF du dossier FACTURE CLIENT
    python pdf_to_excel.py MCI        # uniquement les factures détectées comme MCI
    python pdf_to_excel.py --force    # régénère même si le fichier Excel existe déjà

Le nom du PDF n'a AUCUNE importance (ex : facture.pdf convient).
Chaque PDF est analysé et le fichier de sortie est nommé d'après son CONTENU :
    - Société  -> ligne "Doit : ..."            (1er mot, ex : BSA, MCI)
    - Mois     -> ligne "Mois de prise en charge : Juillet 2026"
    - Année    -> même ligne
Sortie : FACTURE CLIENT/<SOCIETE>/<SOCIETE> <MOIS> <ANNEE>.xlsx
    exemple : FACTURE CLIENT/BSA/BSA JUILLET 2026.xlsx
    (un sous-dossier est créé automatiquement pour chaque société)

Pas de fichier consolidé. Les PDF restent dans le dossier mère FACTURE CLIENT.
Les fichiers Excel déjà existants ne sont PAS écrasés (protection des modifications
manuelles), sauf avec l'option --force.
"""
import re
import sys
import glob
import os
import unicodedata
import pdfplumber
import openpyxl.styles
from openpyxl import Workbook, load_workbook

# Le script est installé directement dans FACTURE CLIENT :
# les PDF et le modèle sont dans son propre dossier.
BASE = os.path.dirname(os.path.abspath(__file__))
PDF_DIR = BASE
MODEL = os.path.join(BASE, "Modele_Import.xlsx")

HEADERS = ["Numero_Facture", "Date_Soins", "Nom_Agent", "Matricule", "Societe",
           "Sous_Societe", "Acte_Medicale_Prix", "Montant_Total_Brut",
           "Ticket_Moderateur", "Prise_En_Charge_Net", "Observations"]

MONTH_ORDER = ["JANVIER", "FEVRIER", "MARS", "AVRIL", "MAI", "JUIN",
               "JUILLET", "AOUT", "SEPTEMBRE", "OCTOBRE", "NOVEMBRE", "DECEMBRE"]


def sans_accent(s):
    """'Février' -> 'FEVRIER' (majuscules sans accent, pour les noms de fichiers)."""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper()


def amount_to_float(s):
    """'1 234,56' -> 1234.56"""
    return float(s.replace("\u00a0", " ").replace(" ", "").replace(",", "."))


def fmt_amount(n):
    """1234.0 -> '1 234' ; 1234.5 -> '1 234,5'"""
    if abs(n - round(n)) < 0.005:
        return f"{int(round(n)):,}".replace(",", " ")
    return f"{n:,.1f}".replace(",", " ").replace(".", ",")


def parse_date(d):
    """'02/01/26' -> '2026-01-02'"""
    dd, mm, yy = d.strip().split("/")
    year = int(yy)
    year += 2000 if year < 100 else 0
    return f"{year:04d}-{mm}-{dd}"


def parse_pdf(path):
    """Extrait (societe, mois, annee, facture, lignes) d'un PDF, tout depuis le CONTENU.

    - societe : 1er mot de la ligne "Doit : ..."       (ex : "MCI CARE" -> MCI)
    - mois/annee : ligne "Mois de prise en charge : Juillet 2026"
    Gère les lignes de continuation : quand une ligne sans N° porte uniquement
    des actes médicaux (coupure de page), ceux-ci sont rattachés à la ligne précédente.
    """
    with pdfplumber.open(path) as pdf:
        facture = doit = mois_txt = None
        annee = ""
        raw_rows = []
        for page in pdf.pages:
            txt = page.extract_text() or ""
            for line in txt.split("\n"):
                m = re.search(r"Facture N°\s*:\s*(\S+)", line)
                if m:
                    facture = m.group(1)
                if doit is None:
                    m = re.search(r"Doit\s*:\s*(.+)", line)
                    if m:
                        doit = m.group(1).strip()
                if mois_txt is None:
                    m = re.search(r"Mois de prise en charge\s*:\s*(.+)", line)
                    if m:
                        mois_txt = unicodedata.normalize("NFC", m.group(1).strip())
            for table in page.extract_tables():
                for row in table:
                    c0 = str(row[0] or "").strip()
                    c4 = str(row[4] or "").strip()
                    if c0 == "N°":
                        continue  # en-tête de tableau
                    if c0 == "Total" or c4 == "Total":
                        continue  # ligne total (le libellé peut être en col 0 ou 4)
                    if not c0:
                        if c4 and raw_rows:
                            # continuation inter-page : rattacher les actes à la ligne précédente
                            prev = raw_rows[-1]
                            prev[4] = (prev[4] or "") + "\n" + c4
                        continue
                    raw_rows.append(list(row))

    # --- Société : 1er mot du "Doit : ..." ; à défaut, 1er mot du nom de fichier ---
    if doit:
        societe = sans_accent(doit.split()[0])
    else:
        societe = sans_accent(os.path.basename(path).split(" ")[0].rsplit(".", 1)[0])
        print(f"!! {os.path.basename(path)} : ligne 'Doit :' introuvable, "
              f"société prise du nom de fichier -> {societe}")

    # --- Mois + année : "Juillet 2026" -> JUILLET / 2026 ---
    mois = annee = ""
    if mois_txt:
        m = re.match(r"^(\S+)\s+(\d{4})$", mois_txt)
        if m:
            mois, annee = sans_accent(m.group(1)), m.group(2)
    if not mois or not annee:
        print(f"!! {os.path.basename(path)} : ligne 'Mois de prise en charge' "
              f"introuvable ou incomplète -> ignoré")
        return societe, mois, annee, facture, []

    lignes = []
    for row in raw_rows:
        num, date, matricule, nom_cell, actes_cell, montant, part, net = row[:8]

        # --- Nom / sous-société ---
        parts = [p.strip() for p in (nom_cell or "").split("\n") if p.strip()]
        sous = ""
        name_parts = []
        for p in parts:
            m = re.fullmatch(r"\((.+)\)", p)
            if m:
                sous = m.group(1).strip()
            else:
                name_parts.append(p)
        nom = " ".join(name_parts)
        nom_agent = f"{nom} ({sous})" if sous else nom

        # --- Actes médicaux ---
        actes = []
        total_actes = 0.0
        for a in (actes_cell or "").split("\n"):
            a = a.strip()
            if not a:
                continue
            m = re.match(r"^(.*?)\s*:\s*([\d\s\u00a0,\.]+)$", a)
            if m:
                lib, val = m.group(1).strip(), amount_to_float(m.group(2))
                total_actes += val
                actes.append(f"{lib} : {fmt_amount(val)}")
            else:
                actes.append(a)
        actes_str = "; ".join(actes)

        brut = amount_to_float(montant)
        tm = amount_to_float(part)
        net_pay = amount_to_float(net)

        # --- Observations ---
        obs = f"Facture mensuelle soins ambulatoires - {mois_txt}"
        if abs(total_actes - brut) > 0.01:
            obs += f" ; ATTENTION : actes détaillés ({fmt_amount(total_actes)} Ar) < montant facturé (écart {fmt_amount(brut - total_actes)} Ar)"

        lignes.append({
            "Numero_Facture": facture,
            "Date_Soins": parse_date(date),
            "Nom_Agent": nom_agent,
            "Matricule": str(matricule or "").strip(),
            "Societe": societe,
            "Sous_Societe": sous,
            "Acte_Medicale_Prix": actes_str,
            "Montant_Total_Brut": int(brut) if brut == int(brut) else brut,
            "Ticket_Moderateur": int(tm) if tm == int(tm) else tm,
            "Prise_En_Charge_Net": int(net_pay) if net_pay == int(net_pay) else net_pay,
            "Observations": obs,
        })
    return societe, mois, annee, facture, lignes


def style_sheet(ws):
    """Applique la mise en forme du modèle (largeurs, police, formats)."""
    model_ws = load_workbook(MODEL)["Modele_Prestations"]
    widths = {k: v.width for k, v in model_ws.column_dimensions.items()}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(name="Calibri", size=12)
    for r in range(2, ws.max_row + 1):
        for c in range(1, 12):
            ws.cell(row=r, column=c).font = openpyxl.styles.Font(name="Calibri", size=12)
        for c in (8, 9, 10):  # montants
            ws.cell(row=r, column=c).number_format = "#,##0"
    ws.freeze_panes = "A2"


def write_workbook(path, lignes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Modele_Prestations"
    ws.append(HEADERS)
    for l in lignes:
        ws.append([l[h] for h in HEADERS])
    style_sheet(ws)
    wb.save(path)


def main():
    args = [a for a in sys.argv[1:] if a != "--force"]
    force = "--force" in sys.argv[1:]
    filtres = [sans_accent(a) for a in args]

    pdfs = sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf")))
    if not pdfs:
        print(f"!! Aucun PDF trouvé dans {PDF_DIR}")
        return
    for pdf_path in pdfs:
        nom_pdf = os.path.basename(pdf_path)
        societe, mois, annee, facture, lignes = parse_pdf(pdf_path)
        if filtres and societe not in filtres:
            continue
        if not lignes:
            print(f"!! {nom_pdf} : aucune ligne trouvée -> ignoré")
            continue
        # sous-dossier de la société, créé automatiquement
        dossier = os.path.join(PDF_DIR, societe)
        os.makedirs(dossier, exist_ok=True)
        out = os.path.join(dossier, f"{societe} {mois} {annee}.xlsx")
        if os.path.exists(out) and not force:
            print(f"-- {societe}\\{societe} {mois} {annee}.xlsx : existe déjà, non écrasé "
                  f"(--force pour régénérer)  [{nom_pdf}]")
            continue
        write_workbook(out, lignes)
        s_brut = sum(l["Montant_Total_Brut"] for l in lignes)
        s_tm = sum(l["Ticket_Moderateur"] for l in lignes)
        s_net = sum(l["Prise_En_Charge_Net"] for l in lignes)
        print(f"OK {societe}\\{societe} {mois} {annee}.xlsx : {facture} | {len(lignes)} lignes | "
              f"Brut {s_brut:,} | TM {s_tm:,} | Net {s_net:,}".replace(",", " ")
              + f"  <- {nom_pdf}")


if __name__ == "__main__":
    main()
